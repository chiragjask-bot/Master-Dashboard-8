import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import re
import io
import os
import random
import zipfile
import gzip
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule

# =====================================================================================
# 0. LOGIN GATE  (keeps the app private without ever putting a password on GitHub)
# =====================================================================================
# Two modes, tried in this order every time the app loads:
#   1) Streamlit secrets [auth] username/password (original method — more secure,
#      since the real credentials never touch this .py file). Used automatically
#      whenever a .streamlit/secrets.toml (or Streamlit Cloud "Secrets") [auth]
#      section is configured.
#   2) A single hardcoded ADMIN_PASSWORD gate (added per request). Used
#      automatically whenever [auth] secrets are NOT configured — e.g. for quick
#      local/admin use without setting up secrets at all.
# .streamlit/secrets.toml, if used, should look like:
#   [auth]
#   username = "your_username"
#   password = "your_password"
ADMIN_PASSWORD = "kano"

# Playful Hinglish error messages for the hardcoded ADMIN_PASSWORD path, one is
# picked at random on every wrong attempt.
ADMIN_PASSWORD_ERROR_MESSAGES = [
    "Password इल्ले! 😅 इल्ले!, खम्मा घणी भाईसा, सॉरी। तुमसे सब कुछ हो पाएगा! यहां बहुत 🤪 दिमाग मत लगाओ, इस वेबसाइट को नहीं, 😂 इस गलत पासवर्ड को छोड़ दो!",
    "❌ Password इल्ले भाईसा! 😅 इल्ले! खम्मा घणी, सॉरी। तुम बाहुबली हो, तुमसे सब कुछ हो पाएगा! पर यहाँ फालतू 🤪 दिमाग मत लगाओ। अपनी सुंदर वेबसाइट को नहीं, 😂 इस सड़े हुए गलत पासवर्ड को छोड़ दो!",
    "❌ खम्मा घणी भाईसा, Password इल्ले! 😅 sorry! तुम तो मंगल ग्रह पर पानी खोज सकते हो, तुमसे सब कुछ हो पाएगा! पर यहाँ ज़्यादा 🤪 दिमाग मत लगाओ। इस सीधे-सादे वेबसाइट को नहीं, 😂 इस जाली पासवर्ड को छोड़ दो!",
    "❌ Password इल्ले! 😅 इल्ले! खम्मा घणी भाईसा, सॉरी। लोड मत लो, तुमसे सब कुछ हो पाएगा! पर यहाँ फालतू 🤪 दिमाग मत लगाओ। दुनिया छोड़ दो, मोक्ष पकड़ लो, पर पहले 😂 इस गलत पासवर्ड को छोड़ दो!",
    "❌ अरे भाईसा! Password इल्ले! 😅 खम्मा घणी, सॉरी। तुम चाहो तो सिस्टम हिला सकते हो, तुमसे सब कुछ हो पाएगा! पर यहाँ ज़्यादा 🤪 दिमाग मत लगाओ। इस निर्दोष वेबसाइट को नहीं, 😂 इस भूतिया गलत पासवर्ड को छोड़ दो!",
]


def check_login():
    if st.session_state.get("authenticated", False):
        return True

    auth_cfg = st.secrets.get("auth", None)

    # ---- Mode 1: secrets-based username/password (unchanged from before) ----
    if auth_cfg:
        st.title("🔒 Financial File Merger & Formatter — Login")
        with st.form("login_form"):
            user = st.text_input("Username")
            pwd = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Log in")
        if submitted:
            if user == auth_cfg.get("username") and pwd == auth_cfg.get("password"):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Invalid username or password.")
        return False

    # ---- Mode 2: hardcoded ADMIN_PASSWORD gate ----
    st.markdown(
        "<p style='text-align: center; margin-top: 100px; color: Green; font-size: 18px;'>"
        "📊 Financial Data File Merger & Formatter</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<h1 style='text-align: center; margin-top: 0px; font-size: 20px;'>🔐 Admin Login</h1>",
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        with st.form("admin_login_form"):
            pwd = st.text_input("Enter Password", type="password")
            submit = st.form_submit_button("Login", use_container_width=True)
            if submit:
                if pwd == ADMIN_PASSWORD:
                    st.session_state["authenticated"] = True
                    st.rerun()
                else:
                    st.error(random.choice(ADMIN_PASSWORD_ERROR_MESSAGES))

    dynamic_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    st.markdown(
        f"<p style='text-align: center; color: gray; font-size: 14px; margin-top: 20px;'>"
        f"Data refreshed: {dynamic_time}</p>",
        unsafe_allow_html=True,
    )
    return False


st.set_page_config(page_title="Financial File Merger & Formatter", layout="wide")

# ==========================================
# 🛡️ HIDE STREAMLIT MENU
# ==========================================
# As supplied, this sets visibility: show — i.e. the hamburger menu, header,
# toolbar, and footer all stay VISIBLE. Flip any of these to "hidden" if you
# actually want that element hidden from viewers.
hide_streamlit_ui = """
<style>
    #MainMenu {visibility: show;}
    header {visibility: show;}
    [data-testid="stToolbar"] {visibility: show;}
    footer {visibility: show;}
</style>
"""
st.markdown(hide_streamlit_ui, unsafe_allow_html=True)

# ==========================================
# 🛡️ HIDE GITHUB ICON ONLY
# ==========================================
hide_github_icon = """
<style>
    [data-testid="stToolbar"] {
        right: 2rem;
    }
    [data-testid="stToolbar"]::before {
        content: "";
    }
    button[kind="header"] {display: none;}
</style>
"""
st.markdown(hide_github_icon, unsafe_allow_html=True)

if not check_login():
    st.stop()

st.title("📊 Financial Data File Merger & Formatter")
st.markdown('<div id="main_tab"></div>', unsafe_allow_html=True)


def jump_to(anchor_id):
    """Sets the scroll target and reruns; the actual scroll happens via the
    scroll-handling snippet placed at the very end of the script."""
    st.session_state["scroll_target"] = anchor_id
    st.rerun()

# =====================================================================================
# 1. Master Sequence Profile Alignment Rule
# =====================================================================================
TAB_SEQUENCE = [
    "EQUITY_L", "SME_EQUITY_L", "Eligible_T0_Securities", "MA", "mcap", "pd", "pr", "bc", "tt",
    "BhavCopy_NSE_CM", "sec_bhavdata_full", "sec_list", "PE", "StocksTraded", "bulk",
    "mrg_trading", "CM_52_wk_High_low", "Pre-Open Market", "eq_band_changes"
]

# Regex parameters handling timestamps, dates, and alphanumeric trailing prefixes
MATCH_PATTERNS = {
    "EQUITY_L": re.compile(r"^EQUITY_L(?:_.*|\d.*)?$", re.IGNORECASE),
    "SME_EQUITY_L": re.compile(r"^SME_EQUITY_L(?:_.*|\d.*)?$", re.IGNORECASE),
    "Eligible_T0_Securities": re.compile(r"^Eligible_T0_Securities(?:_.*|\d.*)?$", re.IGNORECASE),
    "MA": re.compile(r"^MA(?:_.*|\d.*)?$", re.IGNORECASE),
    "mcap": re.compile(r"^mcap(?:_.*|\d.*)?$", re.IGNORECASE),
    "pd": re.compile(r"^pd(?:_.*|\d.*)?$", re.IGNORECASE),
    "pr": re.compile(r"^pr(?:_.*|\d.*)?$", re.IGNORECASE),
    "bc": re.compile(r"^bc(?:_.*|\d.*)?$", re.IGNORECASE),
    "tt": re.compile(r"^tt(?:_.*|\d.*)?$", re.IGNORECASE),
    "BhavCopy_NSE_CM": re.compile(r"^BhavCopy_NSE_CM(?:_.*|\d.*)?$", re.IGNORECASE),
    "sec_bhavdata_full": re.compile(r"^sec_bhavdata_full(?:_.*|\d.*)?$", re.IGNORECASE),
    "sec_list": re.compile(r"^sec_list(?:_.*|\d.*)?$", re.IGNORECASE),
    "PE": re.compile(r"^PE(?:_.*|\d.*)?$", re.IGNORECASE),
    "StocksTraded": re.compile(r"^StocksTraded(?:_.*|\d.*)?$", re.IGNORECASE),
    "bulk": re.compile(r"^bulk(?:_.*|\d.*)?$", re.IGNORECASE),
    "mrg_trading": re.compile(r"^mrg_trading(?:_.*|\d.*)?$", re.IGNORECASE),
    "CM_52_wk_High_low": re.compile(r"^CM_52_wk_High(?:_low)?(?:_.*|\d.*)?$", re.IGNORECASE),
    "Pre-Open Market": re.compile(r"^Pre-Open Market(?:_.*|\d.*)?$", re.IGNORECASE),
    "eq_band_changes": re.compile(r"^eq_band_changes(?:_.*|\d.*)?$", re.IGNORECASE),
}

# 1b. Columns that should be pre-selected (default-checked) for removal on specific tabs.
#     The person can still uncheck any of these in the UI — this only sets the starting state.
DEFAULT_REMOVE_COLUMNS = {
    "Eligible_T0_Securities": [
        "Schedule of list of securities eligible for trading in T+0 Settlement Cycle across Exchanges"
    ],
    "mcap": ["Trade Date", "Last Trade Date"],
    "pd": ["MKT", "IND_SEC", "CORP_IND"],
    "pr": ["MKT"],
    "BhavCopy_NSE_CM": ["TradDt", "BizDt", "FinInstrmTp"],
    "eq_band_changes": ["Sr. No"],
    "Pre-Open Market": ["FFM CAP", "FFM CAP (₹ Crores)"],
}


# 1b-2. Default column *order* per tab, matching the raw layout each source file
#       ships with. This only sets the sequencer's starting order — the person can
#       still drag/move columns afterward. Tabs not listed here just keep whatever
#       order the uploaded file has.
DEFAULT_COLUMN_ORDER = {
    "EQUITY_L": ["SYMBOL", "ISIN NUMBER", "NAME OF COMPANY", "SERIES", "DATE OF LISTING",
                 "MARKET LOT", "PAID UP VALUE", "FACE VALUE"],
    "SME_EQUITY_L": ["SYMBOL", "ISIN_NUMBER", "NAME_OF_COMPANY", "SERIES", "DATE_OF_LISTING",
                     "PAID_UP_VALUE", "FACE_VALUE"],
    "Eligible_T0_Securities": ["Symbol", "Name Of Company", "Series", "Effective Date"],
    "mcap": ["Symbol", "Security Name", "Series", "Face Value(Rs.)", "Market Cap(Rs.)", "Issue Size", "Close Price/Paid up value(Rs.)", "Trade Date", "Last Trade Date"],
    "pd": ["SYMBOL", "SECURITY", "SERIES", "PREV_CL_PR", "OPEN_PRICE", "HIGH_PRICE", "LOW_PRICE",
           "CLOSE_PRICE", "NET_TRDVAL", "NET_TRDQTY", "TRADES", "HI_52_WK", "LO_52_WK"],
    "pr": ["SECURITY", "PREV_CL_PR", "OPEN_PRICE", "HIGH_PRICE", "LOW_PRICE", "CLOSE_PRICE",
           "NET_TRDVAL", "NET_TRDQTY", "TRADES", "HI_52_WK", "LO_52_WK", "IND_SEC", "CORP_IND"],
    "bc": ["SYMBOL", "SECURITY", "SERIES", "PURPOSE", "RECORD_DT", "EX_DT"],
    "tt": ["SECURITY", "NET_TRDVAL", "NET_TRDQTY", "PREV_CL_PR", "CLOSE_PRIC"],
    "BhavCopy_NSE_CM": ["TckrSymb", "ISIN", "FinInstrmNm", "TtlTradgVol", "TtlTrfVal",
                        "TtlNbOfTxsExctd", "PrvsClsgPric", "ClsPric", "LastPric", "OpnPric",
                        "HghPric", "LwPric", "SttlmPric", "NewBrdLotQty", "Sgmt", "FinInstrmId",
                        "Src", "SctySrs", "SsnId"],
    "sec_bhavdata_full": ["SYMBOL", "SERIES", "DELIV_QTY", "DELIV_PER", "PREV_CLOSE",
                          "CLOSE_PRICE", "TTL_TRD_QNTY", "TURNOVER_LACS", "NO_OF_TRADES",
                          "OPEN_PRICE", "HIGH_PRICE", "LOW_PRICE", "LAST_PRICE", "AVG_PRICE", "DATE1"],
    "sec_list": ["Symbol", "Security Name", "Series", "Band", "Remarks"],
    "StocksTraded": ["Symbol", "LTP", "%chng", "Mkt Cap (₹ Crores)", "Volume (Lakhs)",
                      "Value (₹ Crores)", "Series"],
    "bulk": ["Symbol", "Security Name", "Client Name", "Buy/Sell", "Quantity Traded",
             "Trade Price / Wght. Avg. Price", "Date", "Remarks"],
    "CM_52_wk_High_low": ["SYMBOL", "Adjusted_52_Week_High", "52_Week_High_Date",
                          "Adjusted_52_Week_Low", "52_Week_Low_DT", "SERIES"],
    "eq_band_changes": ["Symbol", "Security Name", "Series", "From", "To"],
}


def default_order_for(tab, cols):
    """Reorders this tab's actual columns to match the configured default sequence,
    tolerating case/whitespace differences between the spec and the real header text.
    Any columns not covered by the default sequence keep their original relative
    order and are appended at the end, so nothing is ever silently dropped."""
    wanted = DEFAULT_COLUMN_ORDER.get(tab)
    if not wanted:
        return list(cols)
    lookup = {str(c).strip().lower(): c for c in cols}
    ordered = []
    used = set()
    for w in wanted:
        key = w.strip().lower()
        match = lookup.get(key)
        if match is not None and match not in used:
            ordered.append(match)
            used.add(match)
    for c in cols:
        if c not in used:
            ordered.append(c)
    return ordered


def default_removals_for(tab, cols):
    """Matches the configured default-removal column names against the actual
    column names on this tab, tolerating case/whitespace differences."""
    wanted = DEFAULT_REMOVE_COLUMNS.get(tab, [])
    if not wanted:
        return []
    normalized_wanted = {w.strip().lower() for w in wanted}
    return [c for c in cols if str(c).strip().lower() in normalized_wanted]


# 1c. Default "start cell" for each tab — this is the single source of truth for
#     both (a) the "Start cell for {tab}" data-crop input's default value, and
#     (b) which cell the tab opens/scrolls to in Excel. Anything not listed here
#     defaults to blank / "A1" ("zero zero" — the normal top-left start).
DEFAULT_START_CELLS = {
    "MA": "B201",
    "Eligible_T0_Securities": "B3",
    "mrg_trading": "A11",
}


def get_default_view_cell(tab):
    return DEFAULT_START_CELLS.get(tab, "A1")


# 1d. Row-freeze note: every tab's header actually lands at the same row (row 2 —
#     right after the "⬆️ Main Tab" nav row), regardless of the tab's own start
#     cell, so freezing right after the header row (done generically below) is
#     already correct for every tab. No per-tab override needed.
CUSTOM_FREEZE_ROWS = {}

# 1e. Which cell the "⬆️ Main Tab" jump-back link is written to, per tab.
#     Defaults to A1; a couple of tabs were reported as needing it at B1 instead.
NAV_LINK_CELL_OVERRIDES = {
    "Eligible_T0_Securities": "B1",
    "MA": "B1",
    "mrg_trading": "B1",
}


def get_nav_link_cell(tab):
    return NAV_LINK_CELL_OVERRIDES.get(tab, "A1")


def render_column_sequencer(state_key, current_columns, allow_delete=False, protected=None,
                             label="Column order", default_order=None):
    """Renders a pick + ◀ Move Left / Move Right ▶ (+ optional 🗑 Delete) control.
    Order (and deletions) persist in st.session_state[state_key] across reruns.
    default_order: optional starting sequence (e.g. from DEFAULT_COLUMN_ORDER) used
    only the first time this sequencer initializes; the person can still reorder
    freely afterward. Falls back to current_columns as-is when not provided.
    Returns the ordered list of column names to use for output.
    """
    protected = protected or []

    if state_key not in st.session_state:
        st.session_state[state_key] = list(default_order) if default_order else list(current_columns)
    order = st.session_state[state_key]

    # Keep in sync with the current column set: drop stale entries, append new ones.
    order = [c for c in order if c in current_columns]
    for c in current_columns:
        if c not in order:
            order.append(c)
    st.session_state[state_key] = order

    if not order:
        return order

    st.caption(f"🔀 {label} — pick a column, then move it left/right" + (" or delete it:" if allow_delete else ":"))
    pick_col, left_col, right_col, del_col = st.columns([3, 1, 1, 1])
    with pick_col:
        pick = st.selectbox(label, options=order, key=f"{state_key}_pick", label_visibility="collapsed")
    with left_col:
        if st.button("◀ Left", key=f"{state_key}_left"):
            i = order.index(pick)
            if i > 0:
                order[i - 1], order[i] = order[i], order[i - 1]
                st.session_state[state_key] = order
                st.rerun()
    with right_col:
        if st.button("Right ▶", key=f"{state_key}_right"):
            i = order.index(pick)
            if i < len(order) - 1:
                order[i + 1], order[i] = order[i], order[i + 1]
                st.session_state[state_key] = order
                st.rerun()
    if allow_delete:
        with del_col:
            disabled = pick in protected
            if st.button("🗑 Delete", key=f"{state_key}_del", disabled=disabled,
                         help="This column is required and can't be deleted" if disabled else None):
                order.remove(pick)
                st.session_state[state_key] = order
                st.rerun()

    st.caption(" → ".join(order))

    # ---- Additional sequence box: type the exact order, comma-separated ----
    type_col, apply_col = st.columns([5, 1])
    with type_col:
        typed = st.text_input(
            "Or type the exact sequence here (comma-separated column names), then Apply:",
            key=f"{state_key}_typed",
            placeholder=", ".join(order),
        )
    with apply_col:
        st.write("")
        apply_clicked = st.button("Apply", key=f"{state_key}_apply")

    if apply_clicked:
        typed_names = [t.strip() for t in typed.split(",") if t.strip()]
        if not typed_names:
            st.warning("Type at least one column name before clicking Apply.")
        else:
            lookup = {c.strip().lower(): c for c in order}
            matched, unmatched = [], []
            for name in typed_names:
                actual = lookup.get(name.strip().lower())
                if actual and actual not in matched:
                    matched.append(actual)
                elif not actual:
                    unmatched.append(name)

            if not allow_delete:
                # Non-deletable sequencers: anything left unmentioned is appended
                # at the end rather than dropped, so no data silently disappears.
                for c in order:
                    if c not in matched:
                        matched.append(c)
            else:
                # Deletable sequencers: protected columns are kept even if the
                # person forgot to type them; everything else not typed is dropped.
                for c in protected:
                    if c in order and c not in matched:
                        matched.append(c)

            if unmatched:
                st.warning(f"Not found (ignored): {', '.join(unmatched)}")
            if matched:
                st.session_state[state_key] = matched
                st.rerun()

    return order


# 2. Strict Custom Numeric Formatting Strings Configuration
NUMBER_FORMATS = {
    'price': '#,##0.00',
    'qty': '#,##0',
    'date': 'DD-MMM-YYYY',  # 4-digit year everywhere
    'percent': '0.00"%"',
    'ratio': '0.00',
    'crores': '#,##0.00" Cr"',
    'lakhs': '#,##0.00" L"',
    'number': '0',
    'text': '@'
}


# =====================================================================================
# 2b. Master_Dashboard-8 builder — joins every consolidated tab on Symbol into one
#     wide reference sheet. Runs automatically at the end of "Execute Structural
#     Consolidation", right after all tabs are written into the same workbook.
# =====================================================================================
MASTER_SHEET_NAME = "Master_Dashboard-8"
MASTER_HIGHLIGHT_COLOR = "EAD1DC"
MASTER_HEADER_SCAN_ROWS = 15
MASTER_SYMBOL_ALIASES = ["SYMBOL", "TckrSymb", "Symb", "Symbol"]
# Safety cap: one Data Validation object is created per row for the Symbol
# "box display" quick-view tooltip. Benchmarked at ~6,000 rows in well under a
# second and a tiny file-size increase, so this ceiling is just a sane backstop,
# not a realistic limit for NSE-sized symbol lists.
MAX_BOX_DISPLAY_ROWS = 20000

MASTER_FIELD_MAP = [
    {"label": "Symbol", "sheet": "BhavCopy_NSE_CM", "aliases": ["TckrSymb", "SYMBOL", "Symb"], "format": "text", "isKey": True},
    # Pinned right next to Symbol (not at the far end with the other hyperlink
    # columns) — feature request: a clickable link "dot" visible without
    # scrolling, right beside the symbol name. Freeze panes below is widened
    # to column C to keep both Symbol and this dot on screen while scrolling.
    {"label": "NSE Chart", "sheet": None, "aliases": [], "format": "text"},
    {"label": "ISIN", "sheet": "BhavCopy_NSE_CM", "aliases": ["ISIN", "ISIN NUMBER"], "format": "text"},
    {"label": "Series", "sheet": "BhavCopy_NSE_CM", "aliases": ["SctySrs", "SERIES", "Series", "Srs"], "format": "text"},
    {"label": "Company Name (Capital)", "sheet": "BhavCopy_NSE_CM",
     "aliases": ["FinInstrmNm", "NAME OF COMPANY", "Name Of Company", "Security Name", "SECURITY", "Security",
                 "COMPANY NAME", "COMPANY'S NAME", "Company Name", "Company's Name"], "format": "text"},
    {"label": "Company Name", "sheet": ["EQUITY_L", "SME_EQUITY_L"],
     "aliases": ["NAME OF COMPANY", "Name Of Company", "Security Name", "SECURITY", "Security",
                 "COMPANY NAME", "NAME OF COMPANY", "NAME_OF_COMPANY", "COMPANY'S NAME", "Company Name", "Company's Name"], "format": "text"},
    {"label": "Date of Listing", "sheet": ["EQUITY_L", "SME_EQUITY_L"],
     "aliases": ["DATE OF LISTING", "DATE_OF_LISTING", "LISTING DATE", "LISTING_DATE", "LISTING DT", ], "format": "date"},
    {"label": "Trade Date", "sheet": "BhavCopy_NSE_CM", "aliases": ["TradDt", "Trade Date"], "format": "date"},
    {"label": "Segment", "sheet": "BhavCopy_NSE_CM", "aliases": ["Src"], "format": "text"},
    {"label": "Market Lot", "sheet": "BhavCopy_NSE_CM", "aliases": ["NewBrdLotQty", "MARKET LOT", "Market Lot"], "format": "qty"},
    {"label": "T0 Tag", "sheet": "Eligible_T0_Securities", "aliases": ["SERIES", "SctySrs", "Srs", "Series"], "format": "text"},
    {"label": "Remarks", "sheet": "sec_list", "aliases": ["Remarks"], "format": "text"},
    {"label": "Face Value", "sheet": ["EQUITY_L", "SME_EQUITY_L"],
     "aliases": ["FACE VALUE", "FACE_VALUE", "Face Value(Rs.)"], "format": "price"},
    {"label": "No. of Trades", "sheet": "BhavCopy_NSE_CM", "aliases": ["TtlNbOfTxsExctd", "No. of Trades", "NO OF TRADES", "TRADES", "Trade", "NO_OF_TRADES"], "format": "qty"},
    {"label": "Traded % against Issue Size", "sheet": None, "aliases": [], "format": "percent", "computed": True},
    {"label": "Trades average Deal", "sheet": None, "aliases": [], "format": "price", "computed": True},
    {"label": "Traded Qty", "sheet": "BhavCopy_NSE_CM",
     "aliases": ["TtlTradgVol", "TTL TRD QNTY", "TRADED QUANTITY", "NET_TRDQTY", "Traded Qty", "NET TRD QTY", "NET TRDQTY", "TTL_TRD_QNTY"], "format": "qty"},
    {"label": "Delivery Qty", "sheet": "sec_bhavdata_full",
     "aliases": ["DELIV QTY", "DELIV QUANTITY", "Delivery quantity", "DELIVERY QNTY", "DELIV_QNTY", "DELIV QNTY", "DELIV_QTY"], "format": "qty"},
    {"label": "Turnover (Rs.)", "sheet": "BhavCopy_NSE_CM",
     "aliases": ["TtlTrfVal", "NET_TRDVAL", "NET_TRD_VAL", "NET TRD VAL", "NET TRDVAL", "Turnover (Rs.)", "NET TRADED VALUE", "Net Traded Value", "Traded Value"], "format": "qty"},
    {"label": "Issue Size", "sheet": "mcap", "aliases": ["Issue Size"], "format": "qty"},
    {"label": "Mkt Cap (Rs. Crores)", "sheet": "StocksTraded", "aliases": ["Mkt Cap (Rs Crores)", "Mkt Cap (\u20b9 Crores)", "Market Cap (\u20b9 Crores)"], "format": "crores"},
    {"label": "Market Cap(Rs.)", "sheet": "mcap", "aliases": ["Market Cap(Rs.)", "Mkt Cap(Rs.)", "Market Cap (Rs.)", "Mkt Cap (Rs.)", "Market Cap(Rs)", "Mkt Cap(Rs)", "Market Cap (Rs)", "Mkt Cap (Rs)"], "format": "qty"},
    {"label": "Delivery %", "sheet": "sec_bhavdata_full",
     "aliases": ["DELIV PER", "DELIV %", "delivery percentage", "Delivery Percentage (%)", "DELIV_PER"], "format": "percent"},
    {"label": "Value (Rs. Crores)", "sheet": "StocksTraded", "aliases": ["Value (Rs Crores)", "Value (\u20b9 Crores)"], "format": "crores"},
    {"label": "Value (Rs.)", "sheet": "BhavCopy_NSE_CM", "aliases": ["TtlTrfVal", "NET_TRDVAL"], "format": "qty"},
    {"label": "Volume (Lakhs)", "sheet": "StocksTraded", "aliases": ["Volume (Lakhs)"], "format": "lakhs"},
    {"label": "Volume", "sheet": "BhavCopy_NSE_CM", "aliases": ["TtlTradgVol", "NET_TRDQTY"], "format": "qty"},    
    {"label": "D% against Band", "sheet": None, "aliases": [], "format": "percent", "computed": True},
    {"label": "Band", "sheet": "sec_list", "aliases": ["Band"], "format": "number"},
    {"label": "% Change", "sheet": "StocksTraded", "aliases": ["%chng", "% Change"], "format": "percent"},
    {"label": "Price Change", "sheet": None, "aliases": [], "format": "price_signed", "computed": True},
    {"label": "Close Price", "sheet": "BhavCopy_NSE_CM", "aliases": ["ClsPric", "CLOSE PRICE", "Close Price", "CLOSE_PRICE"], "format": "price"},
    {"label": "CMP/LTP", "sheet": "BhavCopy_NSE_CM", "aliases": ["LastPric", "LAST PRICE", "Last Price", "LTP", "LAST_PRICE"], "format": "price"},
    {"label": "Prev Close", "sheet": "BhavCopy_NSE_CM", "aliases": ["PrvsClsgPric", "PREV CLOSE", "Previous close", "PREV_CL_PR", "PREV_CLOSE"], "format": "price"},
    {"label": "Open (Rs.)", "sheet": "BhavCopy_NSE_CM", "aliases": ["OpnPric", "Open Price", "OPEN PRICE", "OPEN_PRICE"], "format": "price"},
    {"label": "High (Rs.)", "sheet": "BhavCopy_NSE_CM", "aliases": ["HghPric", "HIGH PRICE", "High Price", "HIGH_PRICE"], "format": "price"},
    {"label": "Low (Rs.)", "sheet": "BhavCopy_NSE_CM", "aliases": ["LwPric", "Low Price", "LOW PRICE", "LOW_PRICE"], "format": "price"},
    {"label": "52W High", "sheet": "CM_52_wk_High_low",
     "aliases": ["Adjusted_52_Week_High", "52_Week_High", "52W_High", "52 Week High", "52W High", "HI_52_WK"], "format": "price"},
    {"label": "52W High Date", "sheet": "CM_52_wk_High_low",
     "aliases": ["52_Week_High_Date", "52 Week High Date", "52_Week_High_DT", "52W High Date", "52 W High Date", "52 W High Dt.", "52W High Dt."], "format": "date"},
    {"label": "52W Low", "sheet": "CM_52_wk_High_low",
     "aliases": ["Adjusted_52_Week_Low", "52 Week Low", "52_Week_Low", "52W_Low", "52W Low", "LO_52_WK"], "format": "price"},
    {"label": "52W Low Date", "sheet": "CM_52_wk_High_low",
     "aliases": ["52_Week_Low_DT", "52 Week Low Date", "52 W Low Date", "52 W Low Dt.", "52W Low Dt."], "format": "date"},
    {"label": "Symbol P/E", "sheet": "PE", "aliases": ["SYMBOL P/E", "Symbol P/E"], "format": "ratio"},
    {"label": "Adjusted P/E", "sheet": "PE", "aliases": ["ADJUSTED P/E", "Adjusted P/E"], "format": "ratio"},
    {"label": "T0 Effective Date", "sheet": "Eligible_T0_Securities", "aliases": ["Effective Date"], "format": "date"},
    {"label": "Paid Up Value", "sheet": ["EQUITY_L", "SME_EQUITY_L"],
     "aliases": ["PAID UP VALUE", "PAID_UP_VALUE"], "format": "price"},
    {"label": "Category", "sheet": "mcap", "aliases": ["Category"], "format": "text"},

    # ---- Live-formula columns (feature request: "Create New Column Name and
    # Just paste it below formula in excel"). Unlike the computed columns above,
    # these are written as REAL Excel/Google Sheets formulas (referencing that
    # row's Symbol/CMP/DMA cells), not pre-computed Python values — because they
    # need either live web data (DMAs) or a clickable link (URL columns), neither
    # of which exists in the source bhavcopy files this app reads.
    # IMPORTANT (confirmed by the instruction doc itself): the DMA/Bull-Bear/
    # Difference formulas use GOOGLEFINANCE(), which is a Google Sheets-only
    # function — it will show #NAME? in Excel and only resolve once the file is
    # opened in Google Sheets, exactly as the doc says ("not work in excel ...
    # working properly in google sheet"). See md_write_master_sheet() for the
    # actual formula text written into each row.
    {"label": "CAR Rating", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Difference from 200 DMA", "sheet": None, "aliases": [], "format": "price_signed"},
    {"label": "50 DMA", "sheet": None, "aliases": [], "format": "price"},
    {"label": "100 DMA", "sheet": None, "aliases": [], "format": "price"},
    {"label": "200 DMA", "sheet": None, "aliases": [], "format": "price"},
    {"label": "Bull/Bear Run Output", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Trading View", "sheet": None, "aliases": [], "format": "text"},
    {"label": "History Data", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Chartlink", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Screener", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Marketsmith", "sheet": None, "aliases": [], "format": "text"},
    {"label": "Zerodha", "sheet": None, "aliases": [], "format": "text"},
]

MASTER_NUMBER_FORMATS = {
    "price": "#,##0.00",
    "price_signed": '+#,##0.00;-#,##0.00;0.00',
    "qty": "#,##0",
    "date": "dd-mmm-yyyy",
    "percent": '0.00"%"',
    "ratio": "0.00",
    "crores": '#,##0.00" Cr"',
    "lakhs": '#,##0.00" L"',
    "number": "0",
    "text": "@",
}

# 2b-i. Master_Dashboard-8 calculated columns (feature request: "Create New Column
# Name" with formula-based data, auto-computed from other Master_Dashboard-8
# columns). Each is computed in Python (not as a live Excel formula) so the value
# is always correct and doesn't depend on cell references shifting when columns
# are reordered/hidden. See compute_master_calculated_fields() below.
#   Price Change                 = Close Price − Prev Close
#   Trades average Deal          = (Traded Qty / No. of Trades) × Close Price
#   Traded % against Issue Size  = (Traded Qty × 100) / Issue Size
#   D% against Band              = (% Change × 100) / Band
MASTER_COMPUTED_FIELD_LABELS = {
    "Price Change", "Trades average Deal", "Traded % against Issue Size", "D% against Band",
}


def _md_num(value):
    """Best-effort float conversion for computed-column inputs; returns None for
    blank/unparseable values so the calculated column is simply left blank too,
    instead of raising or writing a wrong 0."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def compute_master_calculated_fields(master_data: dict):
    """Mutates master_data (symbol -> {label: value}) in place, adding the four
    MASTER_COMPUTED_FIELD_LABELS wherever their source columns are available for
    that symbol. Safe to call even if a source column is missing/blank for a
    given row — that computed cell is just left blank rather than erroring."""
    for symbol, rec in master_data.items():
        close = _md_num(rec.get("Close Price"))
        prev = _md_num(rec.get("Prev Close"))
        if close is not None and prev is not None:
            rec["Price Change"] = close - prev

        qty = _md_num(rec.get("Traded Qty"))
        trades = _md_num(rec.get("No. of Trades"))
        if qty is not None and trades not in (None, 0) and close is not None:
            rec["Trades average Deal"] = (qty / trades) * close

        issue_size = _md_num(rec.get("Issue Size"))
        if qty is not None and issue_size not in (None, 0):
            rec["Traded % against Issue Size"] = (qty * 100) / issue_size

        pct_change = _md_num(rec.get("% Change"))
        band = _md_num(rec.get("Band"))
        if pct_change is not None and band not in (None, 0):
            rec["D% against Band"] = (pct_change * 100) / band


# 2c. Column-hide feature (feature request: "add hide column feature in tab name:
#     Master_Dashboard-8"). Fixed list of Master_Dashboard-8 columns that can be
#     hidden/unhidden as a group with a single on/off toggle in the UI. The columns
#     still exist in the sheet and in every download — "hidden" here is openpyxl's
#     column_dimensions[...].hidden flag, exactly like manually hiding a column in
#     Excel (Format -> Hide & Unhide -> Hide Columns). Unhiding in Excel itself
#     always works regardless of this flag.
MASTER_HIDE_COLUMNS = [
    "ISIN",
    "Series",
    "Company Name (Capital)",
    "Company Name",
    "Date of Listing",
    "Trade Date",
    "Segment",
    "T0 Tag",
    "Remarks",
    "Face Value",
    "Market Lot",
    "Mkt Cap (Rs. Crores)",
    "Value (Rs. Crores)",
    "Volume (Lakhs)",
    "Band",
    "Open (Rs.)",
    "High (Rs.)",
    "Low (Rs.)",
    "T0 Effective Date",
    "Symbol P/E",
    "Paid Up Value",
    "Category",
]

# Feature request: "Column hide button size decrease 60%" — scale factor applied
# to the outline +/- toggle's boundary-column width (1.0 = no change; 0.4 = a
# 60% decrease from the sheet's normal 20-wide columns).
MASTER_HIDE_BUTTON_WIDTH_SCALE = 0.4

# 2d. Series-column filter defaults (feature request: "filter data need (required
#     only: EQ, BE, SM, ST) in Tab Name: Master_Dashboard-8"). Both lists are
#     editable in the UI; these are just the starting values in the text boxes.
DEFAULT_SERIES_KEEP = "EQ, BE, SM, ST"
DEFAULT_NAME_EXCLUDE = "ETF, TRUST, REIT, GOLDBONDS, GOI LOAN, GOLD LOAN"


def filter_master_dashboard_rows(df, keep_series_csv="", exclude_name_csv=""):
    """Applies the two Master_Dashboard-8 row filters:
      1) Keep-list on the 'Series' column (exact match, case/space-insensitive).
         Leave the box blank to keep every Series value untouched.
      2) Exclude rows whose Company Name (Capital)/Company Name contains any of
         the given substrings (case-insensitive), e.g. to drop ETF/TRUST/REIT rows.
    Returns a new, reset-index DataFrame; never mutates the input.
    """
    out = df.copy()

    keep_values = {v.strip().upper() for v in keep_series_csv.split(",") if v.strip()}
    if keep_values and "Series" in out.columns:
        out = out[out["Series"].astype(str).str.strip().str.upper().isin(keep_values)]

    exclude_terms = [v.strip().upper() for v in exclude_name_csv.split(",") if v.strip()]
    if exclude_terms:
        name_cols = [c for c in ["Company Name (Capital)", "Company Name"] if c in out.columns]
        if name_cols:
            mask = pd.Series(False, index=out.index)
            for c in name_cols:
                col_upper = out[c].astype(str).str.upper()
                for term in exclude_terms:
                    mask = mask | col_upper.str.contains(re.escape(term), na=False)
            out = out[~mask]

    return out.reset_index(drop=True)


def get_active_master_field_map():
    """MASTER_FIELD_MAP plus any custom columns the person has added via the
    'Add a column from any tab' box in the UI. Custom columns live only in
    st.session_state for this session — nothing is written back to this file."""
    return MASTER_FIELD_MAP + st.session_state.get("custom_master_fields", [])


def md_normalize_header(text) -> str:
    if text is None:
        return ""
    return " ".join(str(text).strip().lower().split())


def build_master_alias_lookup(field_map):
    """Maps each source sheet name -> the set of normalized column-name aliases
    that the given Master_Dashboard-8 field map (MASTER_FIELD_MAP, or the active
    field map including custom columns) pulls from it. Used to highlight, on the
    '⬆️ Main Tab' hub sheet, exactly which columns of each source tab feed into
    Master_Dashboard-8 (feature request: "highlight particular text ... take
    reference of MASTER_FIELD_MAP")."""
    lookup = {}
    for f in field_map:
        sheets = f["sheet"] if isinstance(f["sheet"], list) else [f["sheet"]]
        aliases_norm = {md_normalize_header(a) for a in f["aliases"]}
        for s in sheets:
            lookup.setdefault(s, set()).update(aliases_norm)
    return lookup


def md_find_header_row(ws, all_aliases_norm: set, scan_rows: int = MASTER_HEADER_SCAN_ROWS) -> int:
    max_row = min(scan_rows, ws.max_row)
    max_col = ws.max_column
    if max_row == 0 or max_col == 0:
        return -1
    best_row, best_score = -1, 0
    for r in range(1, max_row + 1):
        score = 0
        for c in range(1, max_col + 1):
            norm = md_normalize_header(ws.cell(row=r, column=c).value)
            if norm and norm in all_aliases_norm:
                score += 1
        if score > best_score:
            best_score, best_row = score, r
    return best_row if best_score > 0 else -1


def md_build_header_index(ws, header_row: int) -> dict:
    idx = {}
    for c in range(1, ws.max_column + 1):
        norm = md_normalize_header(ws.cell(row=header_row, column=c).value)
        if norm:
            idx[norm] = c
    return idx


def md_match_column(header_index: dict, aliases: list) -> int:
    for a in aliases:
        norm = md_normalize_header(a)
        if norm in header_index:
            return header_index[norm]
    return -1


def md_build_master_dashboard(wb, field_map=None):
    """wb is an openpyxl Workbook already holding the freshly consolidated tabs
    (values, not formulas — safe to read cell.value directly, no data_only reload needed).
    field_map: optional field list to use instead of the base MASTER_FIELD_MAP — pass
    get_active_master_field_map() to include any custom columns the user has added."""
    field_map = field_map if field_map is not None else MASTER_FIELD_MAP
    log = []
    all_aliases = list(MASTER_SYMBOL_ALIASES)
    for f in field_map:
        all_aliases += f["aliases"]
    all_aliases_norm = {md_normalize_header(a) for a in all_aliases}

    fields_by_sheet = {}
    for f in field_map:
        if f.get("computed") or f.get("sheet") is None:
            continue  # computed columns (see MASTER_COMPUTED_FIELD_LABELS) have no source sheet
        sheets = f["sheet"] if isinstance(f["sheet"], list) else [f["sheet"]]
        for s in sheets:
            fields_by_sheet.setdefault(s, []).append(f)

    master_data = {}
    symbol_order = []

    for sheet_name, fields in fields_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            log.append(f'Master_Dashboard-8: "{sheet_name}" tab not present in this workbook — skipped.')
            continue
        ws = wb[sheet_name]
        header_row = md_find_header_row(ws, all_aliases_norm)
        if header_row == -1:
            log.append(f'Master_Dashboard-8: could not detect a header row on "{sheet_name}" — skipped.')
            continue
        header_index = md_build_header_index(ws, header_row)
        symbol_col = md_match_column(header_index, MASTER_SYMBOL_ALIASES)
        if symbol_col == -1:
            log.append(f'Master_Dashboard-8: no Symbol-like column found on "{sheet_name}" — skipped.')
            continue
        field_cols = [md_match_column(header_index, f["aliases"]) for f in fields]

        for r in range(header_row + 1, ws.max_row + 1):
            symbol_raw = ws.cell(row=r, column=symbol_col).value
            if symbol_raw is None or str(symbol_raw).strip() == "":
                continue
            symbol = str(symbol_raw).strip()
            if symbol not in master_data:
                master_data[symbol] = {}
                symbol_order.append(symbol)

            for f, col in zip(fields, field_cols):
                label = f["label"]
                if f.get("isKey"):
                    master_data[symbol][label] = symbol
                    continue
                if col == -1:
                    continue
                val = ws.cell(row=r, column=col).value
                cur = master_data[symbol].get(label)
                if cur is None or cur == "":
                    master_data[symbol][label] = val

    compute_master_calculated_fields(master_data)

    symbol_order = sorted(symbol_order)
    labels = [f["label"] for f in field_map]
    rows = [[master_data[s].get(l, "") for l in labels] for s in symbol_order]
    df = pd.DataFrame(rows, columns=labels)
    return df, log


def md_write_master_sheet(wb, df, column_order=None, field_map=None, hide_columns=None):
    """Adds/overwrites Master_Dashboard-8 directly on the same workbook object.
    column_order: optional list of field labels (subset/reordered) controlling
    which columns appear and in what order. Defaults to the full field_map.
    field_map: optional field list to use instead of the base MASTER_FIELD_MAP — pass
    get_active_master_field_map() to include any custom columns the user has added.
    hide_columns: optional iterable of field labels to hide (openpyxl column
    'hidden' flag) on this sheet — pass MASTER_HIDE_COLUMNS when the "hide
    columns" toggle is on, or None/[] to leave every column visible."""
    field_map = field_map if field_map is not None else MASTER_FIELD_MAP
    hide_set = {md_normalize_header(h) for h in (hide_columns or [])}
    if MASTER_SHEET_NAME in wb.sheetnames:
        del wb[MASTER_SHEET_NAME]
    ws = wb.create_sheet(MASTER_SHEET_NAME)

    field_lookup = {f["label"]: f for f in field_map}
    labels = [l for l in (column_order or list(df.columns)) if l in field_lookup]
    if not labels:
        labels = list(df.columns)
    df = df[labels]
    formats = [MASTER_NUMBER_FORMATS.get(field_lookup[l]["format"], "@") for l in labels]

    header_fill = PatternFill(start_color=MASTER_HIGHLIGHT_COLOR, end_color=MASTER_HIGHLIGHT_COLOR, fill_type="solid")
    bold_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            v = None if (val == "" or pd.isna(val)) else val
            # Fix: date-labelled columns sometimes arrive as plain text (e.g.
            # "05-AUG-2026") straight from the source bhavcopy file rather than
            # a real Excel date serial. A text value can carry a date NUMBER
            # FORMAT for display, but Excel's TODAY()-based conditional
            # formatting rules below silently never match text — which is what
            # made the 52W High/Low Date highlighting look "missing". Coercing
            # to a real date here fixes the display AND the conditional format.
            if v is not None and formats[c - 1] == MASTER_NUMBER_FORMATS["date"] and not hasattr(v, "year"):
                parsed = pd.to_datetime(v, errors="coerce", dayfirst=False)
                if pd.notna(parsed):
                    v = parsed.to_pydatetime()
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body_font
            cell.number_format = formats[c - 1]
            cell.border = border

    # ---- Live-formula columns (feature request: "just paste it below formula
    # in excel", DMA/Bull-Bear/Difference/CAR + all 7 hyperlink columns). Written
    # as real formula strings per row, referencing that row's own Symbol/CMP/DMA
    # cells, so they behave like a normal dragged-down Excel/Sheets formula.
    if "Symbol" in labels and len(df) > 0:
        sym_col = ws.cell(row=1, column=labels.index("Symbol") + 1).column_letter
        cmp_col = ws.cell(row=1, column=labels.index("CMP/LTP") + 1).column_letter if "CMP/LTP" in labels else None
        d50_col = ws.cell(row=1, column=labels.index("50 DMA") + 1).column_letter if "50 DMA" in labels else None
        d100_col = ws.cell(row=1, column=labels.index("100 DMA") + 1).column_letter if "100 DMA" in labels else None
        d200_col = ws.cell(row=1, column=labels.index("200 DMA") + 1).column_letter if "200 DMA" in labels else None
        difd200_col = ws.cell(row=1, column=labels.index("Difference from 200 DMA") + 1).column_letter if "Difference from 200 DMA" in labels else None

        # NSE/TradingView/Chartlink/etc. — (url_prefix, url_suffix, link display text).
        # Verified live against each site on 2026-08-05 except Marketsmith (that
        # tool sits behind a login wall with no confirmed public per-symbol URL,
        # so its query-string form here is a best-effort guess, not confirmed).
        # Display text is a small clickable "dot" symbol instead of a full word
        # (feature request: "clickable url hyperlink @ like small dot in quick
        # view ... =HYPERLINK(url,"🟢")") — one colour per link so the columns
        # stay visually distinct at a glance; the column header already says
        # which service it is, so the cell itself only needs to be a compact,
        # clickable dot. Clicking the dot opens the URL exactly like before —
        # only the display text changed, not the link behaviour.
        # Format: "Label": (prefix, suffix, display_prefix, use_cell_in_display)
        HYPERLINK_SPECS = {
            "NSE Chart": (
                "https://www.nseindia.com/get-quotes/equity?symbol=",
                "",
                "🟢",
                False,
            ),  # Static display
            "Trading View": (
                "https://www.tradingview.com/symbols/",
                "",
                "Tre ",
                True,
            ),
            "History Data": (
                "https://www.equitypandit.com/historical-data/",
                "",
                "his ",
                True,
            ),  # Replace with actual base URL
            "Screener": (
                "https://www.screener.in/company/",
                "",
                "Scr ",
                True,
            ),  # Replace with actual base URL structure
            "Zerodha": (
                "https://zerodha.com/markets/stocks/NSE/",
                "",
                "Z ",
                True,
            ),  # Replace with actual base URL
            "Chartlink": (
                "https://www.chartlink.com/stocks/",
                "/evaluation.jsp",
                ".html",
                "CL ",
                True,
            ),  # Replace with actual base URL
            "Marketsmith": (
                "https://marketsmithindia.com/mstool/eval/",
                "/evaluation.jsp",
                "ms ",
                True,
            ),  # Replace with actual base URL
        }

        for r in range(2, len(df) + 2):
            sym_cell = f"{sym_col}{r}"

            for label, (
                prefix,
                suffix,
                display_val,
                use_cell_in_display,
            ) in HYPERLINK_SPECS.items():
                if label in labels:
                    col = labels.index(label) + 1

                    if use_cell_in_display:
                        formula = (
                            f'=HYPERLINK("{prefix}"&{sym_cell}&"{suffix}","{display_val}"&'
                            f"{sym_cell})"
                        )
                    else:
                        formula = (
                            f'=HYPERLINK("{prefix}"&{sym_cell}&"{suffix}","{display_val}")'
                        )

                    ws.cell(row=r, column=col, value=formula)

            # DMAs — GOOGLEFINANCE-based, Google Sheets only (per the doc's own
            # note). Exact formula requested: SORT(GOOGLEFINANCE("NSE:"&Symbol,
            # "price", TODAY()-lookback, TODAY()), 1, 0) — sorts newest-date-first,
            # then QUERY "select Col2 limit N" takes the N most recent closes.
            # Lookback (in calendar days) is wider than the DMA window itself so
            # weekends/holidays don't leave the QUERY short of real trading rows:
            # 50 DMA -> 75d, 100 DMA -> 150d, 200 DMA -> 300d, per the request.
            for label, days, lookback in (("50 DMA", 50, 75), ("100 DMA", 100, 150), ("200 DMA", 200, 300)):
                if label in labels:
                    col = labels.index(label) + 1
                    formula = (
                        f'=IFERROR(AVERAGE(QUERY(SORT(GOOGLEFINANCE("NSE:"&{sym_cell},"price",'
                        f'TODAY()-{lookback},TODAY()),1,0),"select Col2 limit {days}")), "")'
                    )
                    ws.cell(row=r, column=col, value=formula)

            if "Bull/Bear Run Output" in labels and cmp_col and d50_col and d100_col and d200_col:
                col = labels.index("Bull/Bear Run Output") + 1
                c_, e50, e100, e200, e400 = f"{cmp_col}{r}", f"{d50_col}{r}", f"{d100_col}{r}", f"{d200_col}{r}", f"{difd200_col}{r}"
                formula = (
                    f'=IF(AND({c_}>{e50}, {c_}>{e100}, {c_}>{e200}, {e400}>=0.01, {e400}<=10), "🟢 Bull", '
                    f'IF(AND({c_}<{e50}, {c_}<{e100}, {c_}<{e200}, {e400}>=-10, {e400}<=-0.01), "🔴 Bear", "⚪ Unconfirmed"))'
                )
                ws.cell(row=r, column=col, value=formula)

            if "Difference from 200 DMA" in labels and cmp_col and d200_col:
                col = labels.index("Difference from 200 DMA") + 1
                # % difference of CMP from its 200 DMA, per the request:
                # =((CMP-200 DMA)*100)/200 DMA  ->  e.g. =((AF2-AX2)*100)/AX2
                formula = f"=(({cmp_col}{r}-{d200_col}{r})*100)/{d200_col}{r}"
                ws.cell(row=r, column=col, value=formula)

            # CAR Rating — exact LET()-based formula supplied in the request.
            # Pulls the 1-year high date, walks the cumulative-average closes
            # since that high, and checks whether the last 10 cumulative
            # averages were rising for 9 straight comparisons (Google Sheets
            # only: LET/SCAN/CHOOSEROWS/SEQUENCE/LAMBDA are Sheets functions,
            # not classic Excel). Symbol cell substituted in for every "A2".
            if "CAR Rating" in labels:
                col = labels.index("CAR Rating") + 1
                car_template = (
                    '=IFERROR(IF(__S__="","ENTER STOCK",'
                    'LET('
                    'raw_high, GOOGLEFINANCE("NSE:" & __S__, "high", TODAY()-365, TODAY()),'
                    'high_date, IFERROR(TO_DATE(QUERY(raw_high, "SELECT Col1 WHERE Col2 IS NOT NULL '
                    'ORDER BY Col2 DESC LIMIT 1 LABEL Col1 \'\'", 1)), TODAY()-30),'
                    'raw_data, IFERROR(GOOGLEFINANCE("NSE:" & __S__, "close", high_date, TODAY()), '
                    'GOOGLEFINANCE("NSE:" & __S__, "close", TODAY()-10, TODAY())),'
                    'prices, IFERROR(CHOOSEROWS(INDEX(raw_data, 0, 2), SEQUENCE(ROWS(raw_data)-1, 1, 2, 1)), {0}),'
                    'count_rows, ROWS(prices),'
                    'cum_avg, SCAN(0, SEQUENCE(count_rows), LAMBDA(a,n, AVERAGE(CHOOSEROWS(prices, SEQUENCE(n))))),'
                    'last_10, IF(count_rows < 10, {0;0;0;0;0;0;0;0;0;0}, '
                    'CHOOSEROWS(cum_avg, SEQUENCE(10, 1, count_rows - 9, 1))),'
                    'check, SUMPRODUCT(--(CHOOSEROWS(last_10, SEQUENCE(9, 1, 2, 1)) > '
                    'CHOOSEROWS(last_10, SEQUENCE(9, 1, 1, 1)))),'
                    'IF(count_rows < 10, "\u26aa Short History", IF(check = 9, '
                    '"\U0001f7e2 Buy/Average Out", "Avoid/Hold \U0001f534"))'
                    ')), "TICKER NOT FOUND")'
                )
                formula = car_template.replace("__S__", sym_cell)
                ws.cell(row=r, column=col, value=formula)

    # ---- Native Excel hide/unhide button (feature request: "hide/unhide button
    # ⬆️ ... not given in tab Name: Master_Dashboard-8"). openpyxl can't draw a
    # clickable shape/macro button, but Excel's built-in column *grouping* gives a
    # real "+ / -" toggle box that lives right inside the sheet — no macro needed.
    # Every column in MASTER_HIDE_COLUMNS is always put in outline group 1 (whether
    # or not it starts hidden), so the toggle exists in every Master_Dashboard-8
    # sheet. The little "1" box in the sheet's top-left corner (above the row
    # numbers, left of the column letters) collapses/expands EVERY grouped column
    # at once — i.e. exactly the single on/off hide/unhide button that was asked
    # for, even though the grouped columns themselves aren't all next to each other.
    group_norm = {md_normalize_header(h) for h in MASTER_HIDE_COLUMNS}
    ws.sheet_properties.outlinePr.summaryRight = True
    ws.sheet_view.showOutlineSymbols = True

    for c in range(1, len(labels) + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = 20
        label_norm = md_normalize_header(labels[c - 1])
        if label_norm in group_norm:
            ws.column_dimensions[col_letter].outlineLevel = 1
        if label_norm in hide_set:
            ws.column_dimensions[col_letter].hidden = True

    # Mark the "collapsed" flag on the column right after each contiguous grouped
    # run (summaryRight=True puts the +/- box there) so the toggle box shows the
    # correct starting +/- state for that run.
    run_start = None
    for c in range(1, len(labels) + 2):  # +1 sentinel pass to close a run at the end
        in_group = c <= len(labels) and md_normalize_header(labels[c - 1]) in group_norm
        if in_group and run_start is None:
            run_start = c
        elif not in_group and run_start is not None:
            run_hidden = all(
                md_normalize_header(labels[i - 1]) in hide_set for i in range(run_start, c)
            )
            boundary_col = min(c, len(labels))
            boundary_letter = ws.cell(row=1, column=boundary_col).column_letter
            ws.column_dimensions[boundary_letter].collapsed = run_hidden
            # Feature request: "hide button size decrease 60%" — Excel draws the
            # little +/- outline toggle inside this boundary column, so its size
            # IS that column's width. Shrinking the width to 40% (i.e. a 60%
            # decrease) makes the button noticeably smaller/more compact. Note:
            # since this boundary column is also a real, visible data column,
            # this narrows that column's display width too — there's no way in
            # Excel to resize just the +/- box without also resizing the column
            # it sits in.
            ws.column_dimensions[boundary_letter].width = 20 * MASTER_HIDE_BUTTON_WIDTH_SCALE
            run_start = None

    # Freeze the header row + Symbol + the pinned NSE Chart dot column (feature
    # request: "freeze 1st column & 1st row"; widened by one column now that
    # the clickable NSE Chart dot sits right next to Symbol, so the dot stays
    # visible on screen too, not just the Symbol name).
    ws.freeze_panes = "C2" if "NSE Chart" in labels else "B2"

    # ---- AutoFilter enabled by default (feature request: "add filter feature
    # (by default) in tab name: Master_Dashboard-8"). Every column gets a
    # dropdown arrow in the header row the moment the file opens — no manual
    # Data > Filter click needed.
    if labels:
        last_col_letter = ws.cell(row=1, column=len(labels)).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    # ---- Conditional formatting: 52W High/Low Date "days ago" highlight
    # (feature request: "CONDITIONAL FORMATTING FOR 52W HIGH/LOW DATES ... Target
    # Columns: 52W High Date and 52W Low Date"). BUGFIX: a 52-week high/low date
    # is always in the PAST (it already happened), so the rule has to look
    # backward from today (TODAY()-N .. TODAY()) — the previous version checked
    # TODAY() .. TODAY()+N (a future window), which a past date can never satisfy,
    # so the highlight silently never fired. Narrowest window first with
    # stop_if_true=True, so e.g. a date 5 days ago gets ONLY the 7-day green
    # highlight, not also the 15/30/180/365-day colors underneath it.
    MASTER_52W_DATE_CF_RULES = [
        (7, "C6EFCE"),    # green
        (15, "FFEB9C"),   # yellow
        (30, "FFD8A8"),   # orange
        (180, "BDD7EE"),  # blue
        (365, "E1D5E7"),  # purple
    ]
    last_row = ws.max_row
    if last_row >= 2:
        for date_label in ("52W High Date", "52W Low Date"):
            if date_label not in labels:
                continue
            col_letter = ws.cell(row=1, column=labels.index(date_label) + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            for days, color in MASTER_52W_DATE_CF_RULES:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                formula = f'AND({first}<>"",{first}<=TODAY(),{first}>=TODAY()-{days})'
                ws.conditional_formatting.add(
                    rng, FormulaRule(formula=[formula], fill=fill, stopIfTrue=True)
                )

        # ---- Conditional formatting: Date of Listing "anniversary" highlight
        # (feature request: "CONDITIONAL FORMATTING FOR Date of Listing Column,
        # automatic highlight when same date and month near 1 month"). Highlights
        # a listing date whenever this year's (or, near year-end, next year's)
        # month/day anniversary falls within the next 30 days — independent of
        # the listing's actual year.
        if "Date of Listing" in labels:
            col_letter = ws.cell(row=1, column=labels.index("Date of Listing") + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            listing_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            this_year = f"DATE(YEAR(TODAY()),MONTH({first}),DAY({first}))"
            next_year = f"DATE(YEAR(TODAY())+1,MONTH({first}),DAY({first}))"
            formula = (
                f"OR(AND({this_year}>=TODAY(),{this_year}<=TODAY()+30),"
                f"AND({next_year}>=TODAY(),{next_year}<=TODAY()+30))"
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[formula], fill=listing_fill, stopIfTrue=True)
            )

        # ---- Conditional formatting: Price Change (green = price up, red =
        # price down) — the doc flagged this column as missing a highlight rule.
        if "Price Change" in labels:
            col_letter = ws.cell(row=1, column=labels.index("Price Change") + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            up_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            down_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}>0"], fill=up_fill, stopIfTrue=True)
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}<0"], fill=down_fill, stopIfTrue=True)
            )

        # ---- Conditional formatting: Bull/Bear Run Output — green for "Bull",
        # red for "Bear", grey for "Neutral" (the doc flagged this column as
        # missing a highlight rule).
        if "Bull/Bear Run Output" in labels:
            col_letter = ws.cell(row=1, column=labels.index("Bull/Bear Run Output") + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            bull_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            bear_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            neutral_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f'{first}="Bull"'], fill=bull_fill, stopIfTrue=True)
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f'{first}="Bear"'], fill=bear_fill, stopIfTrue=True)
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f'{first}="Neutral"'], fill=neutral_fill, stopIfTrue=True)
            )

        # ---- Conditional formatting: Difference from 200 DMA — green when the
        # CMP is above the 200 DMA (positive), red when it's below (negative)
        # (the doc flagged this column as missing a highlight rule).
        if "Difference from 200 DMA" in labels:
            col_letter = ws.cell(row=1, column=labels.index("Difference from 200 DMA") + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            up_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            down_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}>0"], fill=up_fill, stopIfTrue=True)
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}<0"], fill=down_fill, stopIfTrue=True)
            )

        # ---- Conditional formatting: % Change — green above +3%, red below
        # -3% (feature request: "Column Name: % Change, above 3% and Below 3%
        # give conditional formatting").
        if "% Change" in labels:
            col_letter = ws.cell(row=1, column=labels.index("% Change") + 1).column_letter
            rng = f"{col_letter}2:{col_letter}{last_row}"
            first = f"{col_letter}2"
            up_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            down_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}>3"], fill=up_fill, stopIfTrue=True)
            )
            ws.conditional_formatting.add(
                rng, FormulaRule(formula=[f"{first}<-3"], fill=down_fill, stopIfTrue=True)
            )

        # ---- Conditional formatting: CMP/LTP, Close Price, Prev Close — near
        # the 52-week HIGH (negative for you) vs near the 52-week LOW (positive
        # for you), with a stronger colour when that high/low happened in the
        # last 7 days and a lighter colour when it happened within the last 30
        # days ("this month"). Feature request formulas, adapted to this sheet's
        # real 52W High/Low columns instead of the M2/O2 example cells:
        #   near 52W High: AND(price<>"",52WHigh<>"",price>=52WHigh*0.92)
        #   near 52W Low:  AND(price<>"",52WLow<>"",price<=52WLow*1.08)
        # Narrowest (7-day) rule goes first with stop_if_true=True so a hit in
        # the last 7 days shows only the strong colour, not both.
        high52_col = ws.cell(row=1, column=labels.index("52W High") + 1).column_letter if "52W High" in labels else None
        highdate52_col = ws.cell(row=1, column=labels.index("52W High Date") + 1).column_letter if "52W High Date" in labels else None
        low52_col = ws.cell(row=1, column=labels.index("52W Low") + 1).column_letter if "52W Low" in labels else None
        lowdate52_col = ws.cell(row=1, column=labels.index("52W Low Date") + 1).column_letter if "52W Low Date" in labels else None
        if high52_col and highdate52_col and low52_col and lowdate52_col:
            strong_high_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
            light_high_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
            strong_low_fill = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")
            light_low_fill = PatternFill(start_color="D6F2DE", end_color="D6F2DE", fill_type="solid")
            for price_label in ("CMP/LTP", "Close Price", "Prev Close"):
                if price_label not in labels:
                    continue
                col_letter = ws.cell(row=1, column=labels.index(price_label) + 1).column_letter
                rng = f"{col_letter}2:{col_letter}{last_row}"
                p = f"{col_letter}2"
                h, hd = f"{high52_col}2", f"{highdate52_col}2"
                l, ld = f"{low52_col}2", f"{lowdate52_col}2"
                near_high = f'{p}<>"",{h}<>"",{p}>={h}*0.92'
                near_low = f'{p}<>"",{l}<>"",{p}<={l}*1.08'
                ws.conditional_formatting.add(
                    rng, FormulaRule(
                        formula=[f'AND({near_high},{hd}<>"",{hd}<=TODAY(),{hd}>=TODAY()-7)'],
                        fill=strong_high_fill, stopIfTrue=True,
                    )
                )
                ws.conditional_formatting.add(
                    rng, FormulaRule(
                        formula=[f'AND({near_high},{hd}<>"",{hd}<=TODAY(),{hd}>=TODAY()-30)'],
                        fill=light_high_fill, stopIfTrue=True,
                    )
                )
                ws.conditional_formatting.add(
                    rng, FormulaRule(
                        formula=[f'AND({near_low},{ld}<>"",{ld}<=TODAY(),{ld}>=TODAY()-7)'],
                        fill=strong_low_fill, stopIfTrue=True,
                    )
                )
                ws.conditional_formatting.add(
                    rng, FormulaRule(
                        formula=[f'AND({near_low},{ld}<>"",{ld}<=TODAY(),{ld}>=TODAY()-30)'],
                        fill=light_low_fill, stopIfTrue=True,
                    )
                )


    # Feature (box display & pop-up box both option) in Symbol column"). Two
    # mechanisms are written to the SAME cell so this works in both apps:
    #   1) Excel Data Validation input-message — auto-shows the instant the
    #      cell is selected/clicked. This is an Excel-only UI feature; Google
    #      Sheets silently drops it on import, which is why "Quick View" looked
    #      like it vanished there.
    #   2) A real cell Comment/Note — this DOES survive an .xlsx -> Google
    #      Sheets import (shows as the small black-triangle note in the
    #      cell's corner, same text, hover to view) and also works as a
    #      hover tooltip in Excel, so it's the cross-platform fallback.
    # Excel caps an input message at ~255 characters, so only a condensed set of
    # the most commonly checked fields is shown here. For the FULL record (every
    # single column, dynamically, regardless of column order) see the optional
    # "MasterDashboardPopup" VBA macro described in EXCEL_MACRO_SETUP.md, which
    # shows a full pop-up (MsgBox) card instead — the "both option" the request asked for.
    if "Symbol" in labels and len(df) <= MAX_BOX_DISPLAY_ROWS:
        symbol_col_idx = labels.index("Symbol") + 1
        quick_fields = [l for l in [
            "CMP/LTP", "% Change", "Prev Close", "Delivery %",
            "52W High", "52W Low", "Value (Rs. Crores)", "Turnover (Rs.)",
            "Volume (Lakhs)", "Mkt Cap (Rs. Crores)", "Symbol P/E",
        ] if l in labels]
        for r, row in enumerate(df.itertuples(index=False), start=2):
            row_dict = dict(zip(labels, row))
            parts = []
            for ql in quick_fields:
                v = row_dict.get(ql, "")
                if v == "" or (isinstance(v, float) and pd.isna(v)):
                    continue
                parts.append(f"{ql}: {v}")
            if not parts:
                continue
            prompt_text = "  |  ".join(parts)
            if len(prompt_text) > 253:
                prompt_text = prompt_text[:250] + "..."
            dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True,
                                 showInputMessage=True, showErrorMessage=False)
            dv.promptTitle = "Quick View"
            dv.prompt = prompt_text
            dv.add(ws.cell(row=r, column=symbol_col_idx).coordinate)
            ws.add_data_validation(dv)

            # Cross-platform fallback: same text as a real cell Comment/Note,
            # so it still shows up (as a Google Sheets note) after import.
            symbol_cell = ws.cell(row=r, column=symbol_col_idx)
            comment = Comment(f"Quick View\n{prompt_text}", "Master Dashboard")
            comment.width = 260
            comment.height = 130
            symbol_cell.comment = comment

    # Move it to the front so it's the first thing a reviewer sees, right after Main Tab.
    wb.move_sheet(MASTER_SHEET_NAME, offset=-(len(wb.sheetnames) - 2))
    return wb


# =====================================================================================
# 2c. PDF export — "filter Feature in pdf" (features 5 & 6).
#     IMPORTANT HONEST NOTE: a PDF is a static/print format — it cannot host live
#     Excel-style AutoFilter dropdown arrows the way an .xlsx sheet can (that's an
#     Excel-only interactive feature; no PDF viewer supports it). What we CAN do,
#     and what this does, is generate the PDF from the SAME rows/columns you've
#     already filtered/cleaned above (dedupe, row removal, value exclusion, column
#     order) — i.e. the PDF always mirrors your current filtered view, tab by tab,
#     with a clickable list of tabs at the front acting like a table of contents.
#     For live, changeable filtering, keep using the .xlsx file — every tab there
#     already has a real AutoFilter enabled.
# =====================================================================================
PDF_MAX_ROWS_PER_TAB = 3000  # safety cap so one huge tab can't blow up PDF build time


def build_pdf_bytes(tabs_dict, title="Financial Data Export"):
    """tabs_dict: dict[tab_name] -> DataFrame (already filtered/cleaned upstream).
    Returns a BytesIO of a landscape multi-section PDF: a clickable list of tabs
    up front (acts like a table of contents / tab switcher), then one section per
    tab with its header row repeating on every page."""
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4), title=title,
        leftMargin=16, rightMargin=16, topMargin=16, bottomMargin=16,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 10)]

    story.append(Paragraph("Tabs in this file (click to jump):", styles["Heading2"]))
    for i, (tab, df) in enumerate(tabs_dict.items()):
        story.append(Paragraph(f'<a href="#tab_{i}" color="blue">{tab} — {len(df)} row(s)</a>', styles["Normal"]))
    story.append(PageBreak())

    avail_width = landscape(A4)[0] - 32
    for i, (tab, df) in enumerate(tabs_dict.items()):
        story.append(Paragraph(f'<a name="tab_{i}"/>{tab}', styles["Heading1"]))
        if df is None or df.empty:
            story.append(Paragraph("No rows after filtering.", styles["Normal"]))
        else:
            df_show = df.head(PDF_MAX_ROWS_PER_TAB)
            if len(df) > PDF_MAX_ROWS_PER_TAB:
                story.append(Paragraph(
                    f"Showing first {PDF_MAX_ROWS_PER_TAB:,} of {len(df):,} rows.", styles["Italic"]
                ))
            data = [[str(c) for c in df_show.columns]] + df_show.astype(str).values.tolist()
            col_count = max(len(df_show.columns), 1)
            col_width = max(avail_width / col_count, 18)
            t = Table(data, repeatRows=1, colWidths=col_width)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#EAD1DC")),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(t)
        story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return buf


def resolve_tab_name(filename):
    """Bypasses custom timestamps to cleanly match dynamic files to structural tab sequences."""
    base_name = filename.rsplit('.', 1)[0]
    # strip known compression suffixes that can appear before the real extension
    for comp_ext in ('.csv', '.xls', '.xlsx'):
        if base_name.lower().endswith(comp_ext):
            base_name = base_name[: -len(comp_ext)]
    for tab_name, pattern in MATCH_PATTERNS.items():
        if pattern.match(base_name):
            return tab_name
    return None


def extract_date_from_filename(filename):
    """Pulls a DDMMYY or DDMMYYYY date out of the trailing digits of a filename,
    e.g. bc290726 / bc_290726 / bc_29072026 / PE_300726.csv -> a date object.
    Used to pick the LATEST file automatically when two dated copies of the
    same tab (e.g. bc300726 and bc290726) are uploaded together.
    Returns None if no parseable trailing date is found."""
    base = filename
    for ext in ('.csv.gz', '.csv.zip', '.csv', '.xlsx', '.xls'):
        if base.lower().endswith(ext):
            base = base[: -len(ext)]
            break
    match = re.search(r'(\d{6,8})$', base)
    if not match:
        return None
    digits = match.group(1)
    if len(digits) not in (6, 8):
        return None
    try:
        day = int(digits[0:2])
        month = int(digits[2:4])
        year_digits = digits[4:]
        year = int(year_digits)
        if len(year_digits) == 2:
            year += 2000
        return datetime(year, month, day).date()
    except ValueError:
        return None


def auto_detect_format(col_name, df_sample_series):
    """Categorizes metadata patterns to determine cell styling formats."""
    name_clean = str(col_name).lower().strip()

    if name_clean == 'band':
        return 'number'
    # IMPORTANT: check 'date' keywords BEFORE 'price' keywords. Columns like
    # "52_Week_High_Date" or "52_Week_Low_DT" contain "high"/"low" and would
    # otherwise be misclassified as price columns, which is exactly what was
    # causing the CM_52_wk_High_low date columns to stay as raw 2-digit-year
    # text instead of being converted to proper dates.
    if any(k in name_clean for k in ['date', 'dt', 'time']):
        return 'date'
    elif any(k in name_clean for k in ['price', 'close', 'open', 'high', 'low', 'vwap', 'prev', 'trdval', 'hi_52', 'lo_52']):
        return 'price'
    elif any(k in name_clean for k in ['qty', 'volume', 'shares', 'traded', 'count', 'trades']):
        return 'qty'
    elif any(k in name_clean for k in ['percent', 'pct', 'chg%', 'return', 'yield']):
        return 'percent'
    elif any(k in name_clean for k in ['ratio', 'pe', 'pb', 'dividend']):
        return 'ratio'
    elif 'crore' in name_clean:
        return 'crores'
    elif 'lakh' in name_clean:
        return 'lakhs'
    elif pd.api.types.is_numeric_dtype(df_sample_series):
        return 'number'
    return 'text'


# =====================================================================================
# 3. Generic tabular-file loader
#    Many NSE source files ship with 1-2 "disclaimer" / "effective date" lines above
#    the real header row. Instead of hard-coding this to a single tab
#    (as the previous version did for Eligible_T0_Securities only), we now DETECT
#    the real header row for every file by trying increasing skiprows counts until
#    pandas can parse it as a proper multi-column table.
# =====================================================================================
def load_tabular_file(file_bytes, filename, max_preamble_lines=8):
    """Returns (dataframe, preamble_lines) where preamble_lines is a list of the
    raw text lines that appeared above the real header (or None if there weren't any)."""
    is_csv = filename.lower().endswith('.csv')

    def _read(skip):
        if is_csv:
            return pd.read_csv(io.BytesIO(file_bytes), skiprows=skip)
        else:
            return pd.read_excel(io.BytesIO(file_bytes), skiprows=skip)

    last_err = None
    for skip in range(0, max_preamble_lines + 1):
        try:
            df = _read(skip)
        except Exception as e:
            last_err = e
            continue
        # A real header row should give us more than 1 usable column.
        if df.shape[1] > 1:
            preamble_lines = None
            if skip > 0:
                if is_csv:
                    text = file_bytes.decode('utf-8', errors='replace')
                    preamble_lines = text.splitlines()[:skip]
                else:
                    # For Excel, just remember the raw first-column cell(s) as a header note
                    raw = pd.read_excel(io.BytesIO(file_bytes), nrows=skip, header=None)
                    preamble_lines = [str(x) for x in raw.iloc[:, 0].tolist()]
            return df, preamble_lines
        # single-column parse succeeded but isn't useful yet -> keep trying more skips

    if last_err:
        raise last_err
    raise ValueError(f"Could not detect a valid header row in {filename}")


def parse_start_cell(text):
    """Parses an Excel-style cell reference like 'B9' into (row_number, col_index).
    row_number is 1-indexed (as typed). col_index is 0-indexed (A=0, B=1, ...).
    Returns (None, None) if the text is blank or not a valid cell reference —
    in that case the tab loads normally starting from A1."""
    if not text or not str(text).strip():
        return None, None
    match = re.match(r'^\s*([A-Za-z]+)\s*(\d+)\s*$', str(text).strip())
    if not match:
        return None, None
    col_letters, row_str = match.group(1).upper(), match.group(2)
    row_number = int(row_str)
    col_index = 0
    for ch in col_letters:
        col_index = col_index * 26 + (ord(ch) - ord('A') + 1)
    col_index -= 1  # convert to 0-indexed
    if row_number < 1 or col_index < 0:
        return None, None
    return row_number, col_index


def load_tabular_file_from_cell(file_bytes, filename, row_number, col_index):
    """Loads a file starting from an explicit cell (e.g. B9): rows above it and
    columns to its left are discarded entirely — they never reach the final sheet.
    The given row becomes the header row."""
    is_csv = filename.lower().endswith('.csv')
    skip = row_number - 1
    if is_csv:
        df = pd.read_csv(io.BytesIO(file_bytes), skiprows=skip)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), skiprows=skip)
    if col_index > 0:
        df = df.iloc[:, col_index:]
    return df


def parse_row_selector(text, n_rows):
    """Parses a string like '2,5,10-15' — 1-indexed row numbers as shown in the
    on-screen preview table — into a set of 0-indexed row positions to drop.
    Invalid tokens are ignored; out-of-range numbers are clipped silently."""
    indices = set()
    if not text:
        return indices
    for part in re.split(r'[,\s]+', text.strip()):
        if not part:
            continue
        if '-' in part:
            bounds = part.split('-', 1)
            try:
                start, end = int(bounds[0]), int(bounds[1])
            except ValueError:
                continue
            if start > end:
                start, end = end, start
            for i in range(start, end + 1):
                idx0 = i - 1
                if 0 <= idx0 < n_rows:
                    indices.add(idx0)
        else:
            try:
                i = int(part)
            except ValueError:
                continue
            idx0 = i - 1
            if 0 <= idx0 < n_rows:
                indices.add(idx0)
    return indices


def clean_dataframe(df):
    """Removes fully-empty rows and fully-empty columns."""
    df = df.dropna(axis=0, how='all')
    df = df.dropna(axis=1, how='all')
    # also drop columns that are entirely empty strings / whitespace
    def _all_blank(col):
        return col.astype(str).str.strip().replace({'nan': ''}).eq('').all()
    blank_cols = [c for c in df.columns if _all_blank(df[c])]
    if blank_cols:
        df = df.drop(columns=blank_cols)
    return df.reset_index(drop=True)


# =====================================================================================
# 4. Zip / gzip ingestion
#    Wraps extracted bytes in an object that looks enough like a Streamlit
#    UploadedFile (has .name and .getvalue()) so the rest of the pipeline
#    doesn't need to know whether a file came from a direct upload or a zip.
# =====================================================================================
class InMemoryFile:
    def __init__(self, name, data: bytes):
        self.name = name
        self._data = data

    def getvalue(self):
        return self._data


def extract_all_files(zip_bytes, _depth=0, _max_depth=3):
    """Recursively walks a zip archive, expanding any nested .zip or .gz members,
    and returns a flat list of InMemoryFile objects for every leaf file found."""
    results = []
    if _depth > _max_depth:
        return results
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                name = os.path.basename(info.filename)
                if not name:
                    continue
                data = zf.read(info.filename)
                lower = name.lower()
                if lower.endswith('.zip'):
                    results.extend(extract_all_files(data, _depth + 1, _max_depth))
                elif lower.endswith('.gz'):
                    try:
                        decompressed = gzip.decompress(data)
                        results.append(InMemoryFile(name[:-3], decompressed))
                    except Exception:
                        pass  # skip unreadable gzip members
                else:
                    results.append(InMemoryFile(name, data))
    except zipfile.BadZipFile:
        pass
    return results


# =====================================================================================
# --- UI Setup Layer ---
# =====================================================================================
col_up1, col_up2 = st.columns(2)
with col_up1:
    uploaded_files = st.file_uploader(
        "Upload Source CSV or Excel Files",
        accept_multiple_files=True,
        type=['csv', 'xlsx', 'xls']
    )
with col_up2:
    uploaded_zips = st.file_uploader(
        "…or upload a ZIP / folder-export containing the source files",
        accept_multiple_files=True,
        type=['zip']
    )

all_candidate_files = list(uploaded_files) if uploaded_files else []

if uploaded_zips:
    for z in uploaded_zips:
        extracted = extract_all_files(z.getvalue())
        all_candidate_files.extend(extracted)
    st.caption(
        f"📦 Extracted {sum(len(extract_all_files(z.getvalue())) for z in uploaded_zips)} "
        f"file(s) from {len(uploaded_zips)} zip archive(s), including nested zip/gz files."
    )

st.markdown("---")
st.subheader("📎 Add a custom tab from any Excel/CSV file")
st.caption(
    "Upload any file that doesn't match one of the fixed tab names above and give it "
    "its own tab name. You can add as many of these as you like, rename any of them "
    "at any time, and their columns can be pulled into Master_Dashboard-8 the same "
    "way as the built-in tabs."
)
st.session_state.setdefault("custom_tabs", [])  # list of {"name","filename","bytes"}
st.session_state.setdefault("custom_tab_uploader_key", 0)

new_tab_file = st.file_uploader(
    "File for the new tab",
    type=['csv', 'xlsx', 'xls'],
    key=f"custom_tab_uploader_{st.session_state['custom_tab_uploader_key']}",
)
default_new_name = ""
if new_tab_file is not None:
    default_new_name = re.sub(r'\.(csv|xlsx|xls)$', '', new_tab_file.name, flags=re.IGNORECASE)
new_tab_name_col, new_tab_btn_col = st.columns([3, 1])
with new_tab_name_col:
    new_tab_name = st.text_input(
        "Tab name for this file", value=default_new_name, key="new_custom_tab_name"
    )
with new_tab_btn_col:
    st.write("")
    add_tab_clicked = st.button("➕ Add as new tab", key="add_custom_tab_btn")

if add_tab_clicked:
    name_clean = new_tab_name.strip()
    existing_names = {t.lower() for t in TAB_SEQUENCE} | {
        t["name"].strip().lower() for t in st.session_state["custom_tabs"]
    }
    if new_tab_file is None:
        st.warning("Upload a file before adding it as a tab.")
    elif not name_clean:
        st.warning("Type a tab name before adding.")
    elif name_clean.lower() in existing_names:
        st.warning(f'"{name_clean}" is already a tab name — pick a different one.')
    else:
        st.session_state["custom_tabs"].append({
            "name": name_clean,
            "filename": new_tab_file.name,
            "bytes": new_tab_file.getvalue(),
        })
        # Bump the uploader's key so the widget resets empty, ready for the next file.
        st.session_state["custom_tab_uploader_key"] += 1
        st.success(f'Added tab "{name_clean}" ({new_tab_file.name}).')
        st.rerun()

if st.session_state["custom_tabs"]:
    st.caption("Custom tabs added so far — rename or remove any of them:")
    for i, ct in enumerate(st.session_state["custom_tabs"]):
        rc1, rc2 = st.columns([4, 1])
        with rc1:
            renamed = st.text_input(
                f"Tab name (from {ct['filename']})",
                value=ct["name"],
                key=f"custom_tab_rename_{i}",
            )
            renamed_clean = renamed.strip()
            if renamed_clean and renamed_clean != ct["name"]:
                others = {t.lower() for t in TAB_SEQUENCE} | {
                    t["name"].strip().lower()
                    for j, t in enumerate(st.session_state["custom_tabs"]) if j != i
                }
                if renamed_clean.lower() in others:
                    st.warning(f'"{renamed_clean}" is already used by another tab — not renamed.')
                else:
                    st.session_state["custom_tabs"][i]["name"] = renamed_clean
        with rc2:
            st.write("")
            if st.button("🗑 Remove tab", key=f"custom_tab_remove_{i}"):
                st.session_state["custom_tabs"].pop(i)
                st.rerun()

# Ready-to-use custom tab file map, keyed by (possibly renamed) tab name.
custom_tab_files = {
    ct["name"]: InMemoryFile(ct["filename"], ct["bytes"]) for ct in st.session_state["custom_tabs"]
}
CUSTOM_TAB_NAMES = list(custom_tab_files.keys())
ALL_TABS = TAB_SEQUENCE + CUSTOM_TAB_NAMES

st.markdown("---")

# Active File Verification Status Dashboard
st.subheader("📋 Sequence Checklist & Missing Files Audit")
status_cols = st.columns(3)

# Group every uploaded/extracted file by the tab it matches. A tab can have
# more than one candidate (e.g. bc300726 AND bc290726 uploaded together) —
# we keep all of them so the person can choose, and rank them so the LATEST
# date (parsed from the trailing digits in the filename) comes first.
tab_candidates = {}
for f in all_candidate_files:
    matched_tab = resolve_tab_name(f.name)
    if matched_tab:
        tab_candidates.setdefault(matched_tab, []).append(f)


def _candidate_sort_key(f):
    d = extract_date_from_filename(f.name)
    if d is None:
        return (1, 0)  # no parseable date -> sink to the bottom, keep upload order
    return (0, -d.toordinal())  # dated files: latest date first


for files in tab_candidates.values():
    files.sort(key=_candidate_sort_key)

valid_files_map = {}
# Custom tabs are unambiguous (one file, deliberately added) — register them directly.
valid_files_map.update(custom_tab_files)

for idx, tab in enumerate(TAB_SEQUENCE):
    col_to_use = status_cols[idx % 3]
    if tab in tab_candidates:
        candidates = tab_candidates[tab]
        if len(candidates) > 1:
            labels = []
            for f in candidates:
                d = extract_date_from_filename(f.name)
                date_tag = f" — {d.strftime('%d-%b-%Y')}" if d else ""
                labels.append(f"{f.name}{date_tag}")
            sel_idx = col_to_use.selectbox(
                f"⚠️ {len(candidates)} files found for {tab} — latest chosen by default:",
                options=list(range(len(candidates))),
                format_func=lambda i, _labels=labels: _labels[i],
                key=f"select_{tab}",
            )
            chosen_file = candidates[sel_idx]
        else:
            chosen_file = candidates[0]
        valid_files_map[tab] = chosen_file
        col_to_use.markdown(f"**✅ {tab}** <small style='color:green;'>({chosen_file.name})</small>", unsafe_allow_html=True)
    else:
        col_to_use.markdown(f"**❌ {tab}** — <span style='color:#d9534f; font-weight:bold;'>Missing</span>", unsafe_allow_html=True)

for idx, tab in enumerate(CUSTOM_TAB_NAMES):
    col_to_use = status_cols[(len(TAB_SEQUENCE) + idx) % 3]
    col_to_use.markdown(
        f"**✅ {tab}** <small style='color:green;'>({custom_tab_files[tab].name} — custom tab)</small>",
        unsafe_allow_html=True,
    )

st.markdown("---")

# =====================================================================================
# 5. Main Tab — quick jump navigation to every section, plus a URL reference box
# =====================================================================================
st.subheader("🏠 Main Tab — Quick Navigation")
st.caption(
    "Click a section below to jump straight to it. Every section further down has an "
    "'⬆️ Back to Main Tab' button to return here."
)

nav_cols = st.columns(3)
for idx, tab in enumerate(ALL_TABS):
    nav_col = nav_cols[idx % 3]
    with nav_col:
        if tab in valid_files_map:
            if st.button(f"➡️ {tab}", key=f"jump_{tab}", use_container_width=True):
                jump_to(f"tab_{tab}")
        else:
            st.button(
                f"🚫 {tab}",
                key=f"jump_disabled_{tab}",
                use_container_width=True,
                disabled=True,
                help="Not available — upload a matching file first.",
            )

st.markdown("---")

# --- URL reference list box (editable) ---
st.subheader("🔗 NSE Reference URLs")
st.caption("Default NSE links are pre-loaded below. Add your own URLs with the box underneath.")

DEFAULT_URLS = [
    "https://www.nseindia.com/market-data/stocks-traded",
    "https://www.nseindia.com/all-reports/",
    "https://www.nseindia.com/static/market-data/securities-available-for-trading",
    "https://www.nseindia.com/market-data/live-equity-market",
    "https://www.nseindia.com/market-data/pre-open-market-cm-and-emerge-market",
]

if "custom_urls" not in st.session_state:
    st.session_state["custom_urls"] = DEFAULT_URLS.copy()

url_in_col, url_btn_col = st.columns([5, 1])
with url_in_col:
    new_url = st.text_input(
        "Add a URL",
        key="new_url_input",
        placeholder="https://...",
        label_visibility="collapsed",
    )
with url_btn_col:
    if st.button("➕ Add URL", key="add_url_btn", use_container_width=True):
        candidate = (new_url or "").strip()
        if candidate and candidate not in st.session_state["custom_urls"]:
            st.session_state["custom_urls"].append(candidate)
            st.rerun()

for i, url in enumerate(st.session_state["custom_urls"]):
    row_link_col, row_del_col = st.columns([9, 1])
    with row_link_col:
        st.markdown(f"🔗 [{url}]({url})")
    with row_del_col:
        if st.button("🗑️", key=f"del_url_{i}", help="Remove this URL"):
            st.session_state["custom_urls"].pop(i)
            st.rerun()

st.markdown("---")

if all_candidate_files or custom_tab_files:
    if not valid_files_map:
        st.warning("⚠️ Bypassed all uploaded files. None of the file names match the required target criteria.")
    else:
        columns_to_remove = {}
        processed_dataframes = {}
        special_headers = {}  # tab -> list[str] preamble lines detected above the header

        st.subheader("🛠️ Component Tuning & Data Previews")

        for tab in ALL_TABS:
            if tab in valid_files_map:
                f = valid_files_map[tab]

                # "Zero Zero" start-cell override — some tabs default to a specific
                # start cell as a special case; every other tab defaults to blank
                # (normal A1 / "zero zero" start) unless the person sets one.
                st.session_state.setdefault(f"start_cell_{tab}", DEFAULT_START_CELLS.get(tab, ""))
                start_cell_value = st.session_state.get(f"start_cell_{tab}", "")
                start_row_number, start_col_index = parse_start_cell(start_cell_value)

                try:
                    file_bytes = f.getvalue()
                    if start_row_number is not None:
                        # Manual crop: rows above and columns left of this cell are dropped entirely.
                        df = load_tabular_file_from_cell(file_bytes, f.name, start_row_number, start_col_index)
                    else:
                        df, preamble_lines = load_tabular_file(file_bytes, f.name)
                        if preamble_lines:
                            special_headers[tab] = preamble_lines
                except Exception as e:
                    st.error(f"Failed to cleanly digest workspace segment {tab}: {e}")
                    continue

                # Strip stray whitespace from column names (e.g. sec_bhavdata_full's " SERIES")
                df.columns = [str(c).strip() for c in df.columns]

                # Remove fully-empty rows/columns per request
                df = clean_dataframe(df)

                st.markdown(f'<div id="tab_{tab}"></div>', unsafe_allow_html=True)
                section_expanded = st.session_state.get("scroll_target") == f"tab_{tab}"
                with st.expander(f"Sheet Setup Profile: {tab} ({f.name})", expanded=section_expanded):
                    if st.button("⬆️ Back to Main Tab", key=f"back_{tab}"):
                        jump_to("main_tab")

                    st.text_input(
                        f"Start cell for {tab} (optional — e.g. B9 means data starts at column B, "
                        "row 9; everything above row 9 and left of column B is dropped and won't "
                        "appear in the final file). Leave blank for a normal A1 start.",
                        key=f"start_cell_{tab}"
                    )

                    cols = df.columns.tolist()

                    col_ctrl, row_ctrl = st.columns(2)

                    with col_ctrl:
                        selected_removals = st.multiselect(
                            f"Columns to remove from {tab}:",
                            options=cols,
                            default=default_removals_for(tab, cols),
                            key=f"remove_{tab}"
                        )
                    columns_to_remove[tab] = selected_removals
                    df_after_cols = df.drop(columns=selected_removals)

                    with row_ctrl:
                        row_text = st.text_input(
                            "Rows to remove (row #s shown in the preview, e.g. 2,5,10-15):",
                            key=f"rowsel_{tab}"
                        )

                    dedupe_col, filter_col_ui, filter_val_ui = st.columns([1, 1, 2])
                    with dedupe_col:
                        remove_dupes = st.checkbox("Remove duplicate rows", key=f"dedupe_{tab}")
                    with filter_col_ui:
                        filter_col = st.selectbox(
                            "Remove rows matching a value in:",
                            options=["(none)"] + cols,
                            key=f"filtercol_{tab}"
                        )
                    exclude_values = []
                    if filter_col != "(none)":
                        unique_vals = sorted(
                            df_after_cols[filter_col].dropna().astype(str).unique().tolist()
                        )[:500]
                        with filter_val_ui:
                            exclude_values = st.multiselect(
                                f"Value(s) to remove from '{filter_col}':",
                                options=unique_vals,
                                key=f"filtervals_{tab}"
                            )

                    df_cleaned = df_after_cols.copy()
                    if remove_dupes:
                        df_cleaned = df_cleaned.drop_duplicates()
                    if exclude_values:
                        df_cleaned = df_cleaned[~df_cleaned[filter_col].astype(str).isin(exclude_values)]
                    df_cleaned = df_cleaned.reset_index(drop=True)

                    rows_to_drop = parse_row_selector(row_text, len(df_cleaned))
                    if rows_to_drop:
                        df_cleaned = df_cleaned.drop(index=sorted(rows_to_drop)).reset_index(drop=True)

                    seq_order = render_column_sequencer(
                        f"colorder_{tab}",
                        df_cleaned.columns.tolist(),
                        allow_delete=False,
                        label=f"Column order for {tab}",
                        default_order=default_order_for(tab, df_cleaned.columns.tolist()),
                    )
                    df_cleaned = df_cleaned[seq_order]

                    processed_dataframes[tab] = df_cleaned
                    st.caption(f"Rows: {len(df)} original → {len(df_cleaned)} after cleanup")
                    st.dataframe(df_cleaned.head(10), use_container_width=True)

        # -----------------------------------------------------------------------
        # Optional AI analysis (uses the Anthropic API; needs an API key)
        # -----------------------------------------------------------------------
        st.markdown("---")
        st.subheader("🤖 AI Analysis (optional)")
        st.caption(
            "Generates a short natural-language summary of each tab (row counts, notable "
            "highs/lows, anomalies). Requires an Anthropic API key — get one at "
            "console.anthropic.com. The key is only kept in this browser session; it is "
            "never written to disk or committed to GitHub."
        )
        api_key = st.text_input("Anthropic API key", type="password", key="anthropic_api_key")
        run_ai = st.button("Generate AI Insights")

        if run_ai:
            if not api_key:
                st.warning("Please enter an Anthropic API key first.")
            else:
                try:
                    import anthropic
                except ImportError:
                    st.error(
                        "The `anthropic` package isn't installed. Run `pip install anthropic` "
                        "in this app's environment and reload the page."
                    )
                else:
                    client = anthropic.Anthropic(api_key=api_key)
                    for tab, df_target in processed_dataframes.items():
                        stats_snippet = df_target.describe(include='all').to_string()[:3000]
                        with st.spinner(f"Analyzing {tab}…"):
                            try:
                                resp = client.messages.create(
                                    model="claude-sonnet-4-6",
                                    max_tokens=400,
                                    messages=[{
                                        "role": "user",
                                        "content": (
                                            f"Here are summary statistics for a financial data tab named "
                                            f"'{tab}' with {len(df_target)} rows and columns "
                                            f"{list(df_target.columns)}:\n\n{stats_snippet}\n\n"
                                            "In 3-4 short bullet points, call out anything notable "
                                            "(unusual spreads, missing data, outliers)."
                                        )
                                    }]
                                )
                                text = "".join(
                                    b.text for b in resp.content if getattr(b, "type", "") == "text"
                                )
                                with st.expander(f"AI insights — {tab}"):
                                    st.markdown(text)
                            except Exception as e:
                                st.error(f"AI analysis failed for {tab}: {e}")

        st.markdown("---")
        st.subheader("➕ Add a column from any tab to Master_Dashboard-8")
        st.caption(
            "Master_Dashboard-8's default fields are the list below, but you can also pull "
            "in any extra column from any tab. Pick the tab it lives on, type the exact "
            "column name as it appears in that tab's file, optionally give it a shorter "
            "label for Master_Dashboard-8, then Add. It's session-only — nothing is saved "
            "back to this script."
        )
        add_tab_col, add_name_col, add_label_col = st.columns([2, 2, 2])
        with add_tab_col:
            master_col_tab = st.selectbox("Tab name", options=ALL_TABS, key="custom_field_tab")
        with add_name_col:
            custom_col_name = st.text_input(
                "Column name (exact, from that tab)", key="custom_field_colname"
            )
        with add_label_col:
            custom_label = st.text_input(
                "Label in Master_Dashboard-8 (optional)", key="custom_field_label"
            )
        add_clicked = st.button("Add column", key="custom_field_add")

        if add_clicked:
            col_name_clean = custom_col_name.strip()
            if not col_name_clean:
                st.warning("Type the column name before clicking Add.")
            else:
                label_clean = custom_label.strip() or col_name_clean
                existing_labels = {f["label"] for f in get_active_master_field_map()}
                if label_clean in existing_labels:
                    st.warning(
                        f'"{label_clean}" is already a Master_Dashboard-8 column — '
                        "pick a different label."
                    )
                else:
                    st.session_state.setdefault("custom_master_fields", [])
                    st.session_state["custom_master_fields"].append({
                        "label": label_clean,
                        "sheet": master_col_tab,
                        "aliases": [col_name_clean],
                        "format": "text",
                    })
                    # Stale sequencer order would otherwise hide the new column until
                    # it's dropped and re-synced, so clear it and let it re-append.
                    st.session_state.pop("master_col_order", None)
                    st.success(f'Added "{label_clean}" (from {master_col_tab} → {col_name_clean}).')
                    st.rerun()

        custom_fields = st.session_state.get("custom_master_fields", [])
        if custom_fields:
            st.caption("Custom columns added so far:")
            for i, f in enumerate(custom_fields):
                c1, c2 = st.columns([5, 1])
                c1.write(f'• **{f["label"]}** ← {f["sheet"]} → {f["aliases"][0]}')
                if c2.button("🗑 Remove", key=f"custom_field_remove_{i}"):
                    st.session_state["custom_master_fields"].pop(i)
                    st.session_state.pop("master_col_order", None)
                    st.rerun()

        st.markdown("---")
        st.subheader("🔀 Master_Dashboard-8 — column order & inclusion")
        st.caption(
            "Master_Dashboard-8's columns come from a fixed field list (plus any custom "
            "columns you added above), not from your uploads, so you can sequence them "
            "any time. Move columns left/right to change their order in the final sheet, "
            "or delete ones you don't want. 'Symbol' is the join key and can't be deleted."
        )
        active_field_map = get_active_master_field_map()
        master_col_order = render_column_sequencer(
            "master_col_order",
            [f["label"] for f in active_field_map],
            allow_delete=True,
            protected=["Symbol"],
            label="Master_Dashboard-8 columns",
        )

        st.markdown("---")
        st.subheader("🙈 Master_Dashboard-8 — hide/unhide columns")
        st.caption(
            "One on/off switch hides this fixed list of columns (as Excel column-hide, "
            "not deletion — the data stays in the sheet, just collapsed) in "
            "Master_Dashboard-8: " + ", ".join(MASTER_HIDE_COLUMNS)
        )
        st.checkbox(
            "Hide these columns in Master_Dashboard-8",
            key="master_hide_cols_toggle",
        )

        st.markdown("---")
        st.subheader("🔀 Master_Dashboard-8 — Series filter")
        st.caption(
            "Row filters applied only to Master_Dashboard-8. Leave a box empty to skip "
            "that filter."
        )
        series_filter_col, name_filter_col = st.columns(2)
        with series_filter_col:
            st.text_input(
                "✅ Keep only these Series values (comma-separated)",
                value=st.session_state.get("series_keep_input", DEFAULT_SERIES_KEEP),
                key="series_keep_input",
                help="e.g. EQ, BE, SM, ST — any Series value not in this list is dropped "
                     "from Master_Dashboard-8.",
            )
        with name_filter_col:
            st.text_input(
                "🗑 Delete rows where Company Name contains (comma-separated)",
                value=st.session_state.get("exclude_name_input", DEFAULT_NAME_EXCLUDE),
                key="exclude_name_input",
                help="e.g. ETF, TRUST, REIT — rows whose Company Name / Company Name "
                     "(Capital) contains any of these (case-insensitive) are dropped.",
            )

        st.markdown("---")

        if st.button("🚀 Execute Structural Consolidation", type="primary"):
            output_stream = io.BytesIO()

            NAV_ROW = 1  # excel row 1 on every data sheet is reserved for the "⬆️ Main Tab" jump-back link
            LINK_FONT = Font(color="0563C1", underline="single", bold=True)

            # Computed early (before the Main Tab hub sheet is built below) so the
            # same field map drives both the Main-Tab highlight and the
            # Master_Dashboard-8 build later in this same block.
            exec_field_map = get_active_master_field_map()
            master_alias_lookup = build_master_alias_lookup(exec_field_map)
            MASTER_HIGHLIGHT_MAIN_TAB_COLOR = "FFF2CC"  # light yellow
            main_tab_highlight_fill = PatternFill(
                start_color=MASTER_HIGHLIGHT_MAIN_TAB_COLOR,
                end_color=MASTER_HIGHLIGHT_MAIN_TAB_COLOR,
                fill_type="solid",
            )

            with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
                exportable_tabs = [t for t in ALL_TABS if t in processed_dataframes]

                # --- Build the "Main Tab" hub sheet first, so it opens as sheet #1 ---
                main_ws = writer.book.create_sheet(title="Main Tab")
                main_ws["A1"] = "📊 Master Financial Data — Main Tab"
                main_ws["A1"].font = Font(bold=True, size=14)
                main_ws["A3"] = "Click a tab name below to jump straight to that sheet."
                main_ws["A3"].font = Font(italic=True)
                main_ws["A4"] = "🟡 Highlighted columns below feed into Master_Dashboard-8 (see MASTER_FIELD_MAP)."
                main_ws["A4"].font = Font(italic=True)

                header_row_num = 5
                fixed_headers = ["Tab Name", "Rows", "Columns"]
                max_col_count = max(
                    (len(processed_dataframes[t].columns) for t in exportable_tabs), default=0
                )
                column_headers = [f"Column {i}" for i in range(1, max_col_count + 1)]
                all_headers = fixed_headers + column_headers

                for c_idx, label in enumerate(all_headers, start=1):
                    main_ws.cell(row=header_row_num, column=c_idx, value=label).font = Font(bold=True)

                for i, tab in enumerate(exportable_tabs, start=1):
                    row = header_row_num + i
                    tab_df = processed_dataframes[tab]

                    link_cell = main_ws.cell(row=row, column=1, value=f"➡️ {tab}")
                    link_cell.hyperlink = f"#'{tab}'!A1"
                    link_cell.font = LINK_FONT
                    main_ws.cell(row=row, column=2, value=len(tab_df))
                    main_ws.cell(row=row, column=3, value=len(tab_df.columns))

                    tab_aliases = master_alias_lookup.get(tab, set())
                    for c_idx, col_name in enumerate(tab_df.columns, start=4):
                        col_cell = main_ws.cell(row=row, column=c_idx, value=str(col_name))
                        if md_normalize_header(col_name) in tab_aliases:
                            col_cell.fill = main_tab_highlight_fill
                            col_cell.font = Font(bold=True)

                main_ws.column_dimensions['A'].width = 45
                main_ws.column_dimensions['B'].width = 12
                main_ws.column_dimensions['C'].width = 12
                for c_idx in range(4, 4 + max_col_count):
                    main_ws.column_dimensions[main_ws.cell(row=header_row_num, column=c_idx).column_letter].width = 22

                for tab in ALL_TABS:
                    if tab in processed_dataframes:
                        df_target = processed_dataframes[tab]

                        preamble_offset = len(special_headers.get(tab, []))
                        start_row = NAV_ROW + preamble_offset  # 0-indexed pandas startrow

                        if tab in special_headers:
                            workbook = writer.book
                            worksheet = workbook.create_sheet(title=tab)
                            writer.sheets[tab] = worksheet
                            for i, line in enumerate(special_headers[tab], start=1):
                                worksheet.cell(row=NAV_ROW + i, column=1, value=line)

                        df_target.to_excel(writer, sheet_name=tab, startrow=start_row, index=False)

                        worksheet = writer.sheets[tab]

                        # Row 1: jump-back link to the Main Tab hub sheet.
                        # Most tabs get it at A1; a couple were reported as needing
                        # it at B1 instead (see NAV_LINK_CELL_OVERRIDES).
                        nav_link_cell_ref = get_nav_link_cell(tab)
                        nav_row_num, nav_col_idx = parse_start_cell(nav_link_cell_ref)
                        if nav_row_num is None:
                            nav_row_num, nav_col_idx = NAV_ROW, 0
                        nav_cell = worksheet.cell(row=nav_row_num, column=nav_col_idx + 1, value="⬆️ Main Tab")
                        nav_cell.hyperlink = "#'Main Tab'!A1"
                        nav_cell.font = LINK_FONT

                        header_row = start_row + 1
                        data_start_row = header_row + 1

                        # Loop cells individually to fix numbers and scrub formatting flaws
                        for col_idx, col_name in enumerate(df_target.columns, start=1):
                            fmt_type = auto_detect_format(col_name, df_target[col_name])
                            excel_format = NUMBER_FORMATS.get(fmt_type, '@')

                            for row_idx in range(data_start_row, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                val = cell.value

                                if val is None:
                                    continue

                                val_clean = str(val).strip() if isinstance(val, str) else val

                                if str(col_name).strip().lower() == 'band':
                                    try:
                                        cell.value = int(float(val_clean))
                                        cell.number_format = '0'
                                    except (ValueError, TypeError):
                                        cell.value = val_clean
                                        cell.number_format = '@'

                                elif fmt_type == 'date':
                                    if val_clean in ['-', '', 'NA']:
                                        cell.value = val_clean
                                        cell.number_format = '@'
                                    else:
                                        try:
                                            date_obj = pd.to_datetime(val_clean, dayfirst=True).to_pydatetime()
                                            cell.value = date_obj
                                            cell.number_format = excel_format
                                        except Exception:
                                            cell.value = val_clean
                                            cell.number_format = '@'

                                elif fmt_type in ['price', 'qty', 'percent', 'ratio', 'crores', 'lakhs', 'number']:
                                    try:
                                        numeric_clean = (
                                            str(val_clean).replace(',', '')
                                            if isinstance(val_clean, str) else val_clean
                                        )
                                        cell.value = float(numeric_clean)
                                        cell.number_format = excel_format
                                    except (ValueError, TypeError):
                                        cell.value = val_clean
                                        cell.number_format = '@'

                                else:
                                    try:
                                        numeric_clean = (
                                            str(val_clean).replace(',', '')
                                            if isinstance(val_clean, str) else val_clean
                                        )
                                        cell.value = float(numeric_clean)
                                        cell.number_format = NUMBER_FORMATS['number']
                                    except (ValueError, TypeError):
                                        cell.value = val_clean
                                        cell.number_format = excel_format

                        # Auto-adjust column widths
                        for col_idx, _ in enumerate(df_target.columns, start=1):
                            column_letter = worksheet.cell(row=header_row, column=col_idx).column_letter
                            max_length = 0
                            for row_cell in worksheet[column_letter]:
                                try:
                                    if len(str(row_cell.value)) > max_length:
                                        max_length = len(str(row_cell.value))
                                except Exception:
                                    pass
                            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

                        # Freeze the header row(s) AND the first column (feature request:
                        # "freeze 1st column in all tab"). Freezing right after the header
                        # row is correct for every tab, since the "⬆️ Main Tab" nav row (1)
                        # + header row (2) precede the data on every sheet.
                        last_frozen_row = CUSTOM_FREEZE_ROWS.get(tab, header_row)
                        worksheet.freeze_panes = worksheet.cell(row=last_frozen_row + 1, column=2).coordinate
                        last_col_letter = worksheet.cell(row=header_row, column=len(df_target.columns)).column_letter
                        worksheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{worksheet.max_row}"

                        # Default cell this tab opens/scrolls to in Excel (A1 unless overridden above).
                        view_cell = get_default_view_cell(tab)
                        worksheet.sheet_view.topLeftCell = view_cell
                        if worksheet.sheet_view.selection:
                            worksheet.sheet_view.selection[0].activeCell = view_cell
                            worksheet.sheet_view.selection[0].sqref = view_cell

                # -----------------------------------------------------------------
                # Auto-build Master_Dashboard-8 by default — no extra click needed.
                # Reads straight off writer.book, which already holds every tab
                # just written above, and appends the joined sheet to it.
                # -----------------------------------------------------------------
                master_df, master_log = md_build_master_dashboard(writer.book, field_map=exec_field_map)

                # Row filters: keep-list on Series, exclude-list on Company Name.
                rows_before_filter = len(master_df)
                master_df = filter_master_dashboard_rows(
                    master_df,
                    keep_series_csv=st.session_state.get("series_keep_input", ""),
                    exclude_name_csv=st.session_state.get("exclude_name_input", ""),
                )
                rows_after_filter = len(master_df)

                active_master_order = st.session_state.get(
                    "master_col_order", [f["label"] for f in exec_field_map]
                )
                master_hide_columns = (
                    MASTER_HIDE_COLUMNS if st.session_state.get("master_hide_cols_toggle") else None
                )
                md_write_master_sheet(
                    writer.book, master_df, column_order=active_master_order,
                    field_map=exec_field_map, hide_columns=master_hide_columns,
                )

            # ---------------------------------------------------------------------
            # Feature request: "at time both excel sheet download" — a second,
            # separate workbook containing ONLY the Master_Dashboard-8 sheet, built
            # from the exact same (already-filtered) master_df, so it always matches
            # the combined file above.
            # ---------------------------------------------------------------------
            master_only_wb = Workbook()
            master_only_wb.remove(master_only_wb.active)
            md_write_master_sheet(
                master_only_wb, master_df, column_order=active_master_order,
                field_map=exec_field_map, hide_columns=master_hide_columns,
            )
            master_only_stream = io.BytesIO()
            master_only_wb.save(master_only_stream)

            # IMPORTANT: st.button() only reports True on the exact run it was
            # clicked — any later rerun (e.g. from clicking one of the PDF
            # buttons below) makes this whole "if" block skip entirely again.
            # So the results are stashed in session_state here, and everything
            # that follows (metrics, downloads, PDF export) is rendered from
            # session_state OUTSIDE this button's "if", where it survives reruns.
            st.session_state["consolidation_result"] = {
                "output_bytes": output_stream.getvalue(),
                "master_only_output_bytes": master_only_stream.getvalue(),
                "master_df": master_df,
                "active_master_order": active_master_order,
                "master_log": master_log,
                "processed_dataframes": processed_dataframes,
                "rows_before_filter": rows_before_filter,
                "rows_after_filter": rows_after_filter,
            }
            # Clear any stale PDFs from a previous run so old data can't be re-downloaded.
            st.session_state.pop("all_tabs_pdf_bytes", None)
            st.session_state.pop("master_pdf_bytes", None)

        # ---------------------------------------------------------------------------
        # Everything below reads from session_state, not local variables, so it keeps
        # working across reruns triggered by the PDF buttons (see note above).
        # ---------------------------------------------------------------------------
        result = st.session_state.get("consolidation_result")
        if result:
            st.success("✅ Consolidation and Formatting Complete!")

            master_df = result["master_df"]
            active_master_order = result["active_master_order"]
            master_log = result["master_log"]
            processed_dataframes = result["processed_dataframes"]

            rows_before_filter = result.get("rows_before_filter")
            rows_after_filter = result.get("rows_after_filter")
            if rows_before_filter is not None and rows_after_filter is not None and rows_before_filter != rows_after_filter:
                st.caption(
                    f"🔀 Series/Company-Name filters kept {rows_after_filter} of "
                    f"{rows_before_filter} Master_Dashboard-8 rows."
                )

            if master_log:
                with st.expander(f"⚠️ Master_Dashboard-8: {len(master_log)} warning(s)"):
                    for line in master_log:
                        st.write("- " + line)

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Master_Dashboard-8 Symbols", len(master_df))
            m2.metric("Columns", len(active_master_order))
            if "Symbol" in master_df.columns:
                m3.metric("Duplicate Symbols", int(master_df["Symbol"].duplicated().sum()))
                m4.metric("Blank Symbols", int((master_df["Symbol"].astype(str).str.strip() == "").sum()))
            st.dataframe(master_df[active_master_order].head(20), use_container_width=True)

            ts = datetime.now().strftime('%Y%m%d_%H%M')

            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    label="📥 Download Formatted Master File",
                    data=result["output_bytes"],
                    file_name=f"Master_Financial_Data_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="download_xlsx_btn",
                )
            with dl_col2:
                st.download_button(
                    label="📥 Download Master_Dashboard-8 Only (separate Excel)",
                    data=result["master_only_output_bytes"],
                    file_name=f"Master_Dashboard-8_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_master_only_xlsx_btn",
                )

            st.markdown("---")
            st.subheader("📄 PDF Export")
            st.caption(
                "A PDF can't host live Excel-style AutoFilter dropdown arrows — that's an "
                "Excel-only interactive feature no PDF viewer supports. What these PDFs DO "
                "give you: the exact rows/columns you've already filtered above (dedupe, row "
                "removal, value exclusion, column order), laid out tab by tab with a clickable "
                "list of tabs up front. For live/changeable filtering, keep using the .xlsx file "
                "— every tab there already has a real AutoFilter enabled."
            )
            pdf_col1, pdf_col2 = st.columns(2)
            with pdf_col1:
                if st.button("📄 Build All-Tabs PDF", key="build_all_pdf_btn"):
                    with st.spinner("Building All-Tabs PDF…"):
                        all_tabs_pdf = build_pdf_bytes(processed_dataframes, title="Financial Data — All Tabs")
                        st.session_state["all_tabs_pdf_bytes"] = all_tabs_pdf.getvalue()
                if st.session_state.get("all_tabs_pdf_bytes"):
                    st.download_button(
                        label="📥 Download All-Tabs PDF",
                        data=st.session_state["all_tabs_pdf_bytes"],
                        file_name=f"All_Tabs_{ts}.pdf",
                        mime="application/pdf",
                        key="download_all_pdf_btn",
                    )
            with pdf_col2:
                if st.button("📄 Build Master_Dashboard-8 PDF", key="build_master_pdf_btn"):
                    with st.spinner("Building Master_Dashboard-8 PDF…"):
                        md_pdf = build_pdf_bytes(
                            {MASTER_SHEET_NAME: master_df[active_master_order]},
                            title="Master_Dashboard-8"
                        )
                        st.session_state["master_pdf_bytes"] = md_pdf.getvalue()
                if st.session_state.get("master_pdf_bytes"):
                    st.download_button(
                        label="📥 Download Master_Dashboard-8 PDF",
                        data=st.session_state["master_pdf_bytes"],
                        file_name=f"Master_Dashboard-8_{ts}.pdf",
                        mime="application/pdf",
                        key="download_master_pdf_btn",
                    )

# =====================================================================================
# 6. Scroll-to-anchor execution — runs last so every anchor already exists in the DOM.
#    Consumes and clears the pending scroll target so it only fires once per click.
# =====================================================================================
_scroll_target = st.session_state.get("scroll_target")
if _scroll_target:
    st.session_state["scroll_target"] = None
    components.html(
        f"""
        <script>
            setTimeout(function() {{
                var el = window.parent.document.getElementById("{_scroll_target}");
                if (el) {{ el.scrollIntoView({{behavior: "smooth", block: "start"}}); }}
            }}, 150);
        </script>
        """,
        height=0,
    )
